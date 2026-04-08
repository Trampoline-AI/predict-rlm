"""Custom spreadsheet skill targeting LibreOffice compatibility."""

from pathlib import Path

from predict_rlm import Skill

# Reuse the formula_eval module from the built-in spreadsheet skill
_BUILTIN_MODULES_DIR = (
    Path(__file__).parent.parent
    / ".venv"
    / "lib"
    / "python3.13"
    / "site-packages"
    / "predict_rlm"
    / "skills"
    / "spreadsheet"
    / "modules"
)

# Fallback: find it relative to the installed package
if not _BUILTIN_MODULES_DIR.exists():
    import predict_rlm.skills.spreadsheet.skill as _builtin_skill_mod

    _BUILTIN_MODULES_DIR = Path(_builtin_skill_mod.__file__).parent / "modules"

libreoffice_spreadsheet_skill = Skill(
    name="spreadsheet-libreoffice",
    instructions="""Use openpyxl and pandas to build, transform, and style spreadsheet files (.xlsx).

Output files will be evaluated by opening them in **LibreOffice Calc** and reading the cached cell values. This has critical implications for formulas.

# Formula Compatibility — CRITICAL

## Spill Functions Do NOT Work

LibreOffice supports FILTER, SORT, UNIQUE, etc. as formulas, but **spill results are not persisted** when the file is saved as xlsx. Only the first cell of a spill array gets cached — all other spilled cells read as empty/None. This means any formula that returns an array into multiple cells **will silently lose data**.

**NEVER use these spill/dynamic-array functions in formulas:**
`FILTER`, `SORT`, `UNIQUE`, `SORTBY`, `SEQUENCE`, `RANDARRAY`

**Also avoid these Excel 365-only functions** (evaluate to `#NAME?` in LibreOffice):

| Category | Forbidden Functions |
|---|---|
| Lambda functions | `LET`, `LAMBDA`, `MAP`, `REDUCE`, `SCAN`, `MAKEARRAY`, `BYROW`, `BYCOL` |
| Text functions | `TEXTSPLIT`, `TEXTBEFORE`, `TEXTAFTER`, `VALUETOTEXT`, `ARRAYTOTEXT` |
| Other Excel-only | `STOCKHISTORY`, `IMAGE`, `CHOOSECOLS`, `CHOOSEROWS`, `TAKE`, `DROP`, `EXPAND`, `WRAPCOLS`, `WRAPROWS`, `TOCOL`, `TOROW`, `VSTACK`, `HSTACK` |

## Safe Functions (single-cell, LibreOffice compatible)

These return one value per cell and work correctly:
- **Math**: `SUM`, `SUMIF`, `SUMIFS`, `SUMPRODUCT`, `AVERAGE`, `AVERAGEIF`, `MIN`, `MAX`, `ROUND`, `ABS`, `MOD`, `INT`
- **Lookup**: `VLOOKUP`, `HLOOKUP`, `XLOOKUP`, `XMATCH`, `INDEX`, `MATCH`, `OFFSET`, `INDIRECT`
- **Logic**: `IF`, `IFS`, `AND`, `OR`, `NOT`, `IFERROR`, `IFNA`
- **Count**: `COUNT`, `COUNTA`, `COUNTIF`, `COUNTIFS`, `COUNTBLANK`
- **Text**: `LEFT`, `RIGHT`, `MID`, `LEN`, `TRIM`, `UPPER`, `LOWER`, `CONCATENATE`, `TEXT`, `VALUE`, `FIND`, `SEARCH`, `SUBSTITUTE`, `REPT`
- **Date**: `DATE`, `YEAR`, `MONTH`, `DAY`, `TODAY`, `NOW`, `DATEDIF`, `EDATE`, `EOMONTH`, `WEEKDAY`, `NETWORKDAYS`
- **Other**: `ROW`, `COLUMN`, `ROWS`, `COLUMNS`, `LARGE`, `SMALL`, `RANK`, `PERCENTILE`

## When to Use Python vs Formulas

| Task | Approach |
|---|---|
| **Filtering rows** to populate a sheet | **Python**: read source rows, write matching rows with openpyxl |
| **Sorting data** into a sheet | **Python**: sort in Python, write sorted rows |
| **Listing unique values** into cells | **Python**: deduplicate in Python, write values |
| **Populating a sheet from another sheet** | **Python**: read source, filter/transform, write row by row |
| **Per-cell calculations** (sums, averages, ratios) | **Formula**: `=SUM(...)`, `=A1/B1`, etc. |
| **Per-cell conditional values** | **Formula**: `=IF(...)`, `=IFERROR(...)` |
| **Per-cell lookups** | **Formula**: `=VLOOKUP(...)` or `=INDEX(MATCH(...))` |
| **Conditional formatting** | **openpyxl**: use `openpyxl.formatting.rule` |

**The rule**: if the result fills multiple cells (filtering, sorting, listing), do it in Python and write the values. If the result is one value per cell (sum, lookup, conditional), use a formula.

## Important: Preserve Existing Data Order

When a workbook already has sheets with data that you need to update (e.g., Paid/NotPaid sheets with filtered records), **do not clear and rebuild from scratch**. Instead:
1. Read the existing sheet to understand its current content and ordering
2. Apply the minimum changes needed to satisfy the instruction
3. If you must rebuild, preserve the original row ordering

---

# Spreadsheet Operations Guide

## Output Standards

### Typography

Every generated workbook should default to a clean, widely available typeface — Arial or Calibri work well. Override only when the user specifies a preference or when an existing file already uses a different font family.

### Formula Integrity

Deliver workbooks with **zero** calculation errors. The five error tokens to watch for are `#REF!`, `#DIV/0!`, `#VALUE!`, `#N/A`, and `#NAME?`. Every file must be scanned and cleared of these before delivery. A `#NAME?` error almost always means you used an Excel-only function — replace it with Python logic.

### Respecting Existing Files

When a user provides a template or partially completed workbook, treat its layout, fonts, colors, and naming conventions as authoritative. Mimic the existing patterns rather than overwriting them with defaults from this guide.

---

## Choosing the Right Python Library

Two libraries cover the vast majority of spreadsheet tasks:

**pandas** is ideal when the job is fundamentally about data: reading a file, filtering rows, computing aggregates, pivoting, merging datasets, or exporting a cleaned table.

**openpyxl** is the right pick when the job involves presentation-layer concerns: cell-level formatting, conditional coloring, named ranges, embedded formulas, merged cells, or when you need to surgically edit a workbook without disturbing its existing structure.

For many tasks you'll use both — pandas to wrangle the data, openpyxl to place it into a formatted workbook.

---

## Reading & Analyzing Spreadsheet Data

```python
import pandas as pd

# Pull in the first sheet
df = pd.read_excel("quarterly_data.xlsx")

# Pull in every sheet as a dictionary of DataFrames
sheets = pd.read_excel("quarterly_data.xlsx", sheet_name=None)

# Quick inspection
df.head()
df.info()
df.describe()

# Export after transformations
df.to_excel("cleaned_output.xlsx", index=False)
```

### Tips for robust reads

- Pin column types to avoid silent misinterpretation: `pd.read_excel("f.xlsx", dtype={"account_id": str})`
- Limit memory usage on wide files by selecting only needed columns: `usecols=["Date", "Amount", "Category"]`
- Parse date columns at read time: `parse_dates=["transaction_date"]`

---

## Formulas vs Python — The Rule

Use LibreOffice-compatible Excel formulas for **per-cell calculations** that reference other cells: sums, averages, lookups, conditional logic. These stay live when the spreadsheet is opened.

Use Python to **write data directly** when the task involves filtering, sorting, deduplicating, or populating one sheet from another. Write the values row by row with openpyxl — don't try to use dynamic array formulas.

### Per-cell formulas (good)

```python
# Totals, ratios, conditionals — these work in LibreOffice
ws["F25"] = "=SUM(F3:F24)"
ws["G10"] = "=G8/G6"
ws["H5"] = '=IF(D5>100, "High", "Low")'
ws["B10"] = "=VLOOKUP(A10, Sheet2!A:C, 3, FALSE)"
```

### Filtering/populating sheets (do in Python)

```python
# Instead of =FILTER(...), iterate and write rows directly
from openpyxl import load_workbook

wb = load_workbook("input.xlsx")
src = wb["Source"]
dest = wb["Filtered"]

# Copy headers
for c in range(1, src.max_column + 1):
    dest.cell(1, c).value = src.cell(1, c).value

# Filter and write matching rows
out_row = 2
for r in range(2, src.max_row + 1):
    date_val = src.cell(r, 4).value
    if isinstance(date_val, datetime) and 2019 <= date_val.year <= 2023:
        for c in range(1, src.max_column + 1):
            dest.cell(out_row, c).value = src.cell(r, c).value
        out_row += 1

wb.save("output.xlsx")
```

---

## End-to-End Workflow

1. **Pick your tool.** Data wrangling → pandas. Formatting and formulas → openpyxl. Often both.
2. **Open or create** the workbook.
3. **Populate and style** — write data, insert formulas, apply formatting.
4. **Save** the `.xlsx` file.
5. **Verify formulas** — import `formula_eval` and call `evaluate()` to compute every formula in memory and check for errors. Fix any issues and re-verify until the report comes back clean.

---

## Building a New Workbook

```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Header row
headers = ["Quarter", "Revenue", "COGS", "Gross Profit"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, name="Arial")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center")

# Sample data
ws.append(["Q1", 150000, 90000])
ws.append(["Q2", 175000, 100000])

# Gross profit formula for each row
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=4).value = f"=B{r}-C{r}"

# Totals
total_row = ws.max_row + 1
ws.cell(row=total_row, column=1, value="Total")
ws.cell(row=total_row, column=2).value = f"=SUM(B2:B{total_row - 1})"
ws.cell(row=total_row, column=3).value = f"=SUM(C2:C{total_row - 1})"
ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{total_row - 1})"

# Adjust column widths
for col_letter in ["A", "B", "C", "D"]:
    ws.column_dimensions[col_letter].width = 16

wb.save("quarterly_summary.xlsx")
```

---

## Modifying an Existing Workbook

```python
from openpyxl import load_workbook

wb = load_workbook("quarterly_summary.xlsx")
ws = wb["Summary"]

# Walk through every sheet
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"Processing: {name}, rows={sheet.max_row}")

# Update a value
ws["B2"] = 162000

# Structural changes
ws.insert_rows(3)          # push row 3 down, insert blank
ws.delete_cols(5)          # remove column E

# Add a new tab
notes = wb.create_sheet("Notes")
notes["A1"] = "Model last updated 2026-04-01"

wb.save("quarterly_summary_v2.xlsx")
```

---

## Verifying Formulas with `formula_eval`

openpyxl writes formula strings into cells but does **not** evaluate them. To verify that every formula resolves correctly, use the `formula_eval` module which computes all values in pure Python — no external applications needed.

```python
from formula_eval import evaluate

report = evaluate("quarterly_summary.xlsx")

if report["ok"]:
    print("All formulas clean")
else:
    for token, info in report["breakdown"].items():
        print(f"{token}: {info['count']} error(s)")
        for cell in info["cells"]:
            print(f"  {cell}")
```

You can also inspect individual computed values programmatically:

```python
report = evaluate("quarterly_summary.xlsx")

# Check a specific cell's computed result
assert report["computed"]["Summary!D4"] == 325000
```

### Understanding the report

The `evaluate()` function returns a dictionary with these keys:

| Key | Type | Description |
|---|---|---|
| `ok` | bool | `True` when every formula resolved without an error token |
| `formulas` | int | Total formula cells found in the workbook |
| `errors` | int | How many cells evaluated to an Excel error |
| `breakdown` | dict | Each error token mapped to `{"count": N, "cells": [...]}` |
| `computed` | dict | Every evaluated cell as `"Sheet!Cell": value` — useful for assertions |

When errors appear: fix the offending formulas in openpyxl, save again, and re-run `evaluate()` until `errors` reaches zero. A `#NAME?` error means you used a function LibreOffice doesn't support — replace it with Python logic.

### Ensuring spreadsheet apps recalculate on open

```python
from formula_eval import ensure_recalc_on_open

ensure_recalc_on_open("quarterly_summary.xlsx")
```

This sets the `fullCalcOnLoad` flag in the workbook metadata, guaranteeing that LibreOffice will do a fresh computation pass when the file is opened.

---

## Formula Debugging Checklist

### Before you build

- Spot-check two or three cell references to make sure they resolve to the values you expect.
- Confirm your column-number-to-letter mapping — column 64 is `BL`, not `BK`.
- Account for the index offset between pandas (0-based) and Excel (1-based). A DataFrame's row index 5 lands on Excel row 6.

### Common traps

- **`#NAME?` errors**: you used an Excel-only function. Replace with Python logic or a LibreOffice-compatible formula.
- **Null values**: screen with `pd.notna()` before writing to cells.
- **Wide datasets**: fiscal-year data frequently sits beyond column 50 — don't stop scanning early.
- **Duplicate matches**: when searching for a label, verify you've found the right occurrence if it appears more than once.
- **Division by zero**: guard every divisor — wrap in `IFERROR` or check the denominator cell.
- **Broken references**: after inserting or deleting rows/columns, confirm existing formulas still point where they should.
- **Multi-sheet links**: the correct syntax is `SheetName!A1`; forgetting the sheet name prefix silently targets the active sheet.

### Incremental testing strategy

- Wire up formulas in a handful of cells first. Recalculate and verify before copying the pattern across hundreds of rows.
- Make sure every cell that a formula depends on actually exists and contains the expected type of data.
- Stress-test with boundary values: zero, negative numbers, very large figures.

---

## openpyxl Essentials

- Rows and columns are **1-indexed**: `cell(row=1, column=1)` is `A1`.
- To read previously computed values without formulas, open with `data_only=True`. **Caution**: saving a workbook opened this way permanently replaces formulas with their last-calculated values.
- For very large files, `read_only=True` (reading) and `write_only=True` (writing) dramatically reduce memory consumption.
- Formulas written by openpyxl are stored as text — use `formula_eval.evaluate()` to verify their computed values in Python.

## pandas Essentials

- Force column types at load time to prevent silent coercion: `dtype={"zip_code": str}`.
- Narrow wide files by selecting only relevant columns: `usecols=["A", "D", "F"]`.
- Let pandas handle date parsing: `parse_dates=["order_date"]`.

---

## Code Style Notes

When writing Python for spreadsheet tasks, favor brevity: short variable names, minimal inline commentary, no diagnostic print statements unless you're actively debugging. The code is a means to an end — the workbook is the deliverable.

Inside the workbook itself, be generous with documentation: annotate complex formulas with cell comments, tag every hardcoded input with its source, and add section labels so future readers can navigate the model without reverse-engineering it.
""",
    packages=["openpyxl", "pandas", "formulas"],
    modules={"formula_eval": str(_BUILTIN_MODULES_DIR / "formula_eval.py")},
)
