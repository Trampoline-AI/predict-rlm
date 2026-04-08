import os
import json
import datetime
import openpyxl
import argparse
import numpy as np
from tqdm import tqdm
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):

    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        # print(type(v1), type(v2))
        return False
    if v1 == v2:
        return True
    else:
        return False


def _get_color_rgb(color) -> str:
    """Extract RGB value from color object, defaulting to '00000000' if not a string."""
    if color and isinstance(color.rgb, str):
        return color.rgb
    return "00000000"


def _compare_colors(color1, color2) -> bool:
    """Compare two colors using only last 6 characters (RGB), ignoring alpha channel."""
    rgb1 = _get_color_rgb(color1)
    rgb2 = _get_color_rgb(color2)
    return rgb1[-6:] == rgb2[-6:]


def compare_fill_color(fill1, fill2) -> bool:
    """Compare fill colors between two cells."""
    return _compare_colors(fill1.fgColor, fill2.fgColor) and _compare_colors(
        fill1.bgColor, fill2.bgColor
    )


def compare_font_color(font_gt, font_proc) -> bool:
    """Compare font colors between two cells."""
    return _compare_colors(font_gt.color, font_proc.color)


def col_num2name(n):
    """ Convert a column number to an Excel column name """
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """ Convert an Excel column name to a column number """
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def parse_cell_range(range_str):
    """ Parse a range string like 'A1:AB12' """
    start_cell, end_cell = range_str.split(':')
    start_col, start_row = '', ''
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char
    
    end_col, end_row = '', ''
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    """ Generate a list of all cell names in the specified range """
    if ':' not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]
    return cell_names


def expand_column_range(cell_range, ws):
    """Expand whole-column ranges like 'A:G' to 'A1:G{max_row}' using the worksheet."""
    if ':' not in cell_range:
        return cell_range
    start, end = cell_range.split(':')
    start_has_row = any(c.isdigit() for c in start)
    end_has_row = any(c.isdigit() for c in end)
    if not start_has_row and not end_has_row:
        return f"{start}1:{end}{ws.max_row}"
    return cell_range


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    if sheet_name not in wb_proc:
        return False, f"Sheet '{sheet_name}' not found in output (has: {wb_proc.sheetnames})"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_range = expand_column_range(cell_range, ws_gt)
    cell_names = generate_cell_names(cell_range)
    diffs = []
    matched = 0

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            diffs.append(f"  {sheet_name}!{cell_name}: expected={cell_gt.value!r} got={cell_proc.value!r}")
        else:
            matched += 1

    total = len(cell_names)
    if diffs:
        header = f"Sheet '{sheet_name}' range {cell_range}: {matched}/{total} cells match"
        return False, header + "\n" + "\n".join(diffs)

    return True, f"Sheet '{sheet_name}' range {cell_range}: all {total} cells match"


def split_answer_position(answer_position):
    """Split answer_position on commas, respecting single-quoted sheet names."""
    # If quotes are balanced, use quote-aware splitting; otherwise fall back to
    # naive split (handles malformed data like Sheet3'!A:G,'Sheet4'!A:G).
    if answer_position.count("'") % 2 != 0:
        return answer_position.split(',')

    parts = []
    current = []
    in_quotes = False
    for ch in answer_position:
        if ch == "'":
            in_quotes = not in_quotes
            current.append(ch)
        elif ch == ',' and not in_quotes:
            parts.append(''.join(current))
            current = []
        else:
            current.append(ch)
    if current:
        parts.append(''.join(current))
    return parts


def compare_workbooks(gt_file, proc_file, instruction_type, answer_position):
    if not os.path.exists(proc_file):
        return False, "Output file does not exist"
    # Open workbooks
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, f"Failed to load workbooks: {e}"

    sheet_cell_ranges = split_answer_position(answer_position)
    result_list = []
    msg_list = []
    for sheet_cell_range in sheet_cell_ranges:
        if '!' in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split('!')
            sheet_name = sheet_name.lstrip("'").rstrip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range

        # process sheet_name and cell_range
        sheet_name = sheet_name.lstrip("'").rstrip("'")
        cell_range = cell_range.lstrip("'").rstrip("'")

        result, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        result_list.append(result)
        msg_list.append(msg)

    return all(result_list), "\n".join(msg_list)


def parse_option():
    parser = argparse.ArgumentParser("command line arguments for evaluation.")
    
    parser.add_argument('--model', type=str, default='llama', help='model name')
    parser.add_argument('--setting', type=str, default='single',
        help='four setting: single, multi_react_exec, multi_row_exec, multi_row_react_exec')
    parser.add_argument('--dataset', type=str, default="all_data_912", help='dataset name')
    parser.add_argument('--log_dir', type=str, default=None,
        help='run log directory to write per-test-case eval logs into (e.g. logs/run_20260407_141500)')
    parser.add_argument('--concurrency', type=int, default=4,
        help='Number of tasks to evaluate in parallel (default: 4)')

    opt = parser.parse_args()

    # Default to latest run dir if not specified
    if opt.log_dir is None:
        latest_file = os.path.join('logs', 'latest')
        if os.path.exists(latest_file):
            with open(latest_file) as f:
                opt.log_dir = f.read().strip()
            print(f"Using latest run: {opt.log_dir}")

    return opt


def evaluate_task(data, dataset_path, output_dir, log_dir):
    """Evaluate a single task. Returns (result_dict, skipped) tuple."""
    task_id = data['id']
    spreadsheet_dir = f"{dataset_path}/spreadsheet/{task_id}"

    # Detect how many test cases exist for this task (1 or 3)
    num_cases = 0
    for idx in range(1, 4):
        gt = f"{spreadsheet_dir}/{idx}_{task_id}_golden.xlsx"
        if not os.path.exists(gt):
            gt = f"{spreadsheet_dir}/{idx}_{task_id}_answer.xlsx"
        if os.path.exists(gt):
            num_cases = idx
        else:
            break
    if num_cases == 0:
        return None, True

    # Check if any output exists for this task
    has_output = any(
        os.path.exists(f"{output_dir}/{idx}_{task_id}_output.xlsx")
        for idx in range(1, num_cases + 1)
    )
    if not has_output:
        return None, True

    # Build answer_position with sheet info when provided separately
    answer_position = data['answer_position']
    answer_sheet = data.get('answer_sheet')
    if answer_sheet and '!' not in answer_position:
        sheets = [s.strip().strip("'") for s in answer_sheet.split(',')]
        ranges = [r.strip() for r in answer_position.split(',')]
        if len(sheets) > 1 and len(ranges) == 1:
            answer_position = ','.join(f"'{s}'!{ranges[0]}" for s in sheets)
        elif len(sheets) == 1:
            answer_position = f"'{sheets[0]}'!{answer_position}"

    test_case_results = []
    for idx in range(1, num_cases + 1):
        gt_path = f"{spreadsheet_dir}/{idx}_{task_id}_golden.xlsx"
        if not os.path.exists(gt_path):
            gt_path = f"{spreadsheet_dir}/{idx}_{task_id}_answer.xlsx"
        proc_path = f"{output_dir}/{idx}_{task_id}_output.xlsx"
        try:
            result, msg = compare_workbooks(gt_path, proc_path, data['instruction_type'], answer_position)
        except Exception as e:
            result, msg = False, f"Exception: {e}"

        if log_dir:
            task_dir = os.path.join(log_dir, str(task_id))
            os.makedirs(task_dir, exist_ok=True)
            eval_log = os.path.join(task_dir, f"eval_{idx}.log")
            with open(eval_log, 'w') as lf:
                status = "PASS" if result else "FAIL"
                lf.write(f"{status}\n")
                lf.write(f"task: {task_id} test_case: {idx}\n")
                lf.write(f"type: {data['instruction_type']}\n")
                lf.write(f"answer_position: {answer_position}\n")
                lf.write(f"gt:   {gt_path}\n")
                lf.write(f"proc: {proc_path}\n")
                lf.write(f"\n{msg}\n")

        test_case_results.append(int(result))

    soft_restriction = test_case_results.count(1) / len(test_case_results)
    hard_restriction = 0 if 0 in test_case_results else 1
    return {
        'id': task_id,
        'instruction_type': data['instruction_type'],
        'test_case_results': test_case_results,
        'soft_restriction': soft_restriction,
        'hard_restriction': hard_restriction,
    }, False


def evaluation(opt):
    dataset_path = os.path.abspath(f'data/{opt.dataset}')
    with open(f'{dataset_path}/dataset.json', 'r') as fp:
        dataset = json.load(fp)

    output_dir = f"{dataset_path}/outputs/{opt.setting}_{opt.model}"

    eval_results = []
    skipped = 0

    with ThreadPoolExecutor(max_workers=opt.concurrency) as pool:
        futures = {
            pool.submit(evaluate_task, data, dataset_path, output_dir, opt.log_dir): data
            for data in dataset
        }
        for future in tqdm(as_completed(futures), total=len(futures)):
            result, was_skipped = future.result()
            if was_skipped:
                skipped += 1
            else:
                eval_results.append(result)

    # Print summary
    total = len(eval_results)
    if total > 0:
        avg_soft = sum(r['soft_restriction'] for r in eval_results) / total
        avg_hard = sum(r['hard_restriction'] for r in eval_results) / total
        print(f"\nEvaluated: {total} tasks (skipped {skipped} without outputs)")
        print(f"Soft restriction (avg): {avg_soft:.4f}")
        print(f"Hard restriction (avg): {avg_hard:.4f}")
        print(f"Tasks with all cases passing: {sum(r['hard_restriction'] for r in eval_results)}/{total}")
    else:
        print(f"\nNo tasks with outputs found in {output_dir}")

    os.makedirs('outputs', exist_ok=True)
    with open(f'outputs/eval_{opt.setting}_{opt.model}.json', 'w') as fp:
        json.dump(eval_results, fp, indent=4)


if __name__ == "__main__":
    opt = parse_option()
    print(opt)

    evaluation(opt)
