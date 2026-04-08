"""
Open and recalculate spreadsheet files so that formula values are cached,
enabling openpyxl to read computed values with data_only=True.

Cross-platform: uses LibreOffice (macOS/Linux) or win32com+Excel (Windows).

Usage:
    python open_spreadsheet.py --dir_path /path/to/spreadsheets

    # To force a specific backend:
    python open_spreadsheet.py --dir_path /path/to/spreadsheets --backend libreoffice
    python open_spreadsheet.py --dir_path /path/to/spreadsheets --backend win32com

    # To specify a custom LibreOffice path:
    python open_spreadsheet.py --dir_path /path/to/spreadsheets --soffice /path/to/soffice

Requirements (one of):
    - LibreOffice 7.5+ (macOS: brew install --cask libreoffice, Linux: apt install libreoffice-calc)
    - Windows with Microsoft Excel and pywin32 installed (pip install pywin32)
"""

import os
import sys
import shutil
import argparse
import subprocess
import tempfile
import platform

import openpyxl


# ---- LibreOffice backend (macOS / Linux / Windows) ----

def find_libreoffice():
    """Locate the LibreOffice binary."""
    candidates = [
        "libreoffice",
        "soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",  # macOS
    ]
    for candidate in candidates:
        path = shutil.which(candidate)
        if path:
            return path
        if os.path.isfile(candidate):
            return candidate
    return None


def just_open_libreoffice(filename, soffice_path):
    """
    Open an xlsx/xls file in LibreOffice Calc headlessly, which triggers
    formula recalculation, then save it back as xlsx so that openpyxl
    can read the cached computed values with data_only=True.

    Uses --convert-to to force a full re-export (which caches computed
    values), then restores original sheet names if LibreOffice renamed any.
    """
    filename = os.path.abspath(filename)
    if not os.path.isfile(filename):
        print(f"File not found: {filename}")
        return False

    original_dir = os.path.dirname(filename)
    basename = os.path.basename(filename)
    name, ext = os.path.splitext(basename)

    # Record original sheet names before conversion
    try:
        orig_wb = openpyxl.load_workbook(filename, read_only=True)
        orig_names = orig_wb.sheetnames
        orig_wb.close()
    except Exception:
        orig_names = None

    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            profile_dir = os.path.join(tmpdir, "profile")
            result = subprocess.run(
                [
                    soffice_path,
                    "--headless",
                    "--calc",
                    "--convert-to", "xlsx:Calc MS Excel 2007 XML",
                    "--outdir", tmpdir,
                    filename,
                ],
                capture_output=True,
                text=True,
                timeout=120,
            )

            if result.returncode != 0:
                print(f"LibreOffice error for {basename}: {result.stderr.strip()}")
                return False

            converted = os.path.join(tmpdir, name + ".xlsx")
            if not os.path.isfile(converted):
                print(f"Converted file not found for {basename}")
                return False

            dest = os.path.join(original_dir, name + ".xlsx")
            shutil.move(converted, dest)

            # If the original was .xls (not .xlsx), remove the old file
            if ext.lower() == ".xls":
                os.remove(filename)

            # Restore sheet names if LibreOffice renamed any during conversion
            if orig_names:
                try:
                    wb = openpyxl.load_workbook(dest)
                    if wb.sheetnames != orig_names and len(wb.sheetnames) == len(orig_names):
                        for ws, orig in zip(wb.worksheets, orig_names):
                            if ws.title != orig:
                                ws.title = orig
                        wb.save(dest)
                    wb.close()
                except Exception:
                    pass

            return True

        except subprocess.TimeoutExpired:
            print(f"Timeout processing: {basename}")
            return False
        except Exception as e:
            print(f"Error processing {basename}: {e}")
            return False


# ---- win32com backend (Windows only) ----

def just_open_win32com(filename):
    """Open and save a spreadsheet using Excel COM automation (Windows only)."""
    from win32com.client import Dispatch

    filename = os.path.abspath(filename)
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlApp.DisplayAlerts = False
    xlApp.ScreenUpdating = False
    try:
        xlBook = xlApp.Workbooks.Open(Filename=filename, UpdateLinks=False, ReadOnly=False)
        xlBook.Save()
        xlBook.Close(SaveChanges=True)
        return True
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
    finally:
        xlApp.Quit()


# ---- Shared logic ----

def open_all_spreadsheet_in_dir(dir_path, backend, soffice_path=None, recursive=False):
    if not os.path.isdir(dir_path):
        print(f"Not a valid dir path: {dir_path}")
        return

    supported_extensions = {'.xlsx', '.xls'}

    if recursive:
        files = sorted(
            os.path.join(root, f)
            for root, _, filenames in os.walk(dir_path)
            for f in filenames
            if os.path.splitext(f)[1].lower() in supported_extensions
        )
    else:
        files = sorted(
            os.path.join(dir_path, f)
            for f in os.listdir(dir_path)
            if os.path.splitext(f)[1].lower() in supported_extensions
        )

    if not files:
        print(f"No spreadsheet files found in {dir_path}")
        return

    print(f"Found {len(files)} spreadsheet(s) to process using {backend} backend")
    success = 0
    failed = 0

    for full_path in files:
        print(f"Processing: {full_path}")
        if backend == "libreoffice":
            ok = just_open_libreoffice(full_path, soffice_path)
        else:
            ok = just_open_win32com(full_path)
        if ok:
            success += 1
        else:
            failed += 1

    print(f"\nDone. {success} succeeded, {failed} failed out of {len(files)} files.")


def detect_backend():
    """Auto-detect the best available backend."""
    if platform.system() == "Windows":
        try:
            from win32com.client import Dispatch
            return "win32com"
        except ImportError:
            pass
    soffice = find_libreoffice()
    if soffice:
        return "libreoffice"
    if platform.system() == "Windows":
        return "win32com"  # will fail with import error, but gives a clear message
    return None


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description="Open and recalculate spreadsheets to cache formula values."
    )
    parser.add_argument('--dir_path', type=str, required=True,
                        help='Directory containing spreadsheet files')
    parser.add_argument('--backend', type=str, choices=['libreoffice', 'win32com'],
                        default=None,
                        help='Recalculation backend (auto-detected if omitted)')
    parser.add_argument('--soffice', type=str, default=None,
                        help='Path to LibreOffice binary (auto-detected if omitted)')
    parser.add_argument('--recursive', action='store_true',
                        help='Recursively process spreadsheets in subdirectories')
    opt = parser.parse_args()

    backend = opt.backend or detect_backend()
    if backend is None:
        print(
            "ERROR: No recalculation backend found.\n"
            "  Install LibreOffice (macOS: brew install --cask libreoffice, "
            "Linux: apt install libreoffice-calc)\n"
            "  Or on Windows, install pywin32: pip install pywin32"
        )
        sys.exit(1)

    soffice_path = None
    if backend == "libreoffice":
        soffice_path = opt.soffice or find_libreoffice()
        if not soffice_path:
            print(
                "ERROR: LibreOffice not found. Install it or pass --soffice /path/to/soffice\n"
                "  macOS:  brew install --cask libreoffice\n"
                "  Linux:  sudo apt install libreoffice-calc"
            )
            sys.exit(1)
        print(f"Using LibreOffice: {soffice_path}")

    open_all_spreadsheet_in_dir(opt.dir_path, backend, soffice_path, recursive=opt.recursive)
