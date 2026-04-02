"""Spreadsheet skill — build, transform, and style .xlsx workbooks in the sandbox."""

from pathlib import Path

from predict_rlm import Skill

_MODULES_DIR = Path(__file__).parent / "modules"

spreadsheet_skill = Skill(
    name="spreadsheet",
    instructions="""Use openpyxl and pandas to build, transform, and style spreadsheet files (.xlsx).

# Spreadsheet Operations Guide

## Output Standards

### Typography

Every generated workbook should default to a clean, widely available typeface — Arial or Calibri work well. Override only when the user specifies a preference or when an existing file already uses a different font family.

### Formula Integrity

Deliver workbooks with **zero** calculation errors. The five error tokens to watch for are `#REF!`, `#DIV/0!`, `#VALUE!`, `#N/A`, and `#NAME?`. Every file must be scanned and cleared of these before delivery.

### Respecting Existing Files

When a user provides a template or partially completed workbook, treat its layout, fonts, colors, and naming conventions as authoritative. Mimic the existing patterns rather than overwriting them with defaults from this guide.

---

## Finance-Specific Conventions

### Cell Color Scheme

Finance professionals expect a specific visual language in models. Apply these unless the user or an existing template dictates otherwise:

| Color | Meaning |
|---|---|
| Blue font `(0, 0, 255)` | Manually entered inputs — the numbers an analyst would change when running scenarios |
| Black font `(0, 0, 0)` | Computed cells — anything driven by a formula |
| Green font `(0, 128, 0)` | Cross-sheet references within the same workbook |
| Red font `(255, 0, 0)` | References that point outside the current file |
| Yellow fill `(255, 255, 0)` | Cells flagged for review — key assumptions or placeholders awaiting real data |

### Numeric Display Rules

Consistent number formatting makes models readable at a glance:

- **Calendar years** — render as plain text so "2026" doesn't become "2,026"
- **Dollar amounts** — use `$#,##0` and always clarify units in the header (e.g., "EBITDA ($K)" or "Revenue ($mm)")
- **Zero values** — display as a dash. A format string like `$#,##0;($#,##0);"-"` handles positive, negative, and zero cases
- **Percentages** — one decimal place by default: `0.0%`
- **Valuation multiples** — show as `0.0x` (e.g., 8.5x for an EV/EBITDA ratio)
- **Negative figures** — wrap in parentheses `(1,200)` rather than prefixing with a minus sign

### Structuring Assumptions

Every driver — growth rates, margin percentages, discount rates, exit multiples — should live in its own dedicated cell. Formulas should reference that cell instead of embedding the number directly. For instance, write `=D10*(1+$C$3)` where `$C$3` holds the growth rate, rather than `=D10*1.07`.

### Guarding Against Formula Errors

Before delivering a model: confirm that every cell reference resolves correctly, watch for off-by-one mistakes in ranges, verify that each projection period uses a structurally identical formula, stress-test with zeros and negatives, and ensure no accidental circular dependencies exist.

### Citing Data Sources

Any manually entered figure should carry a source annotation — either as a cell comment or in an adjacent column at the row's end. Use a consistent pattern:

```
Ref: [Source Name], [Filing/Report], [Date], [Page/Section], [Link if available]
```

Examples:
- `Ref: Annual Report FY2025, p. 52, Revenue Breakdown, https://...`
- `Ref: Bloomberg, MSFT US Equity, pulled 2025-09-01`
- `Ref: S&P Capital IQ, Consensus Estimates, retrieved 2025-08-20`

---

## Choosing the Right Python Library

Two libraries cover the vast majority of spreadsheet tasks:

**pandas** is ideal when the job is fundamentally about data: reading a file, filtering rows, computing aggregates, pivoting, merging datasets, or exporting a cleaned table.

**openpyxl** is the right pick when the job involves presentation-layer concerns: cell-level formatting, conditional coloring, named ranges, embedded formulas, merged cells, or when you need to surgically edit a workbook without disturbing its existing structure.

For many tasks you'll use both — pandas to wrangle the data, openpyxl to place it into a formatted workbook.

---

## Reading & Analyzing Spreadsheet Data

```python
import pandas as pd

# Pull in the first sheet
df = pd.read_excel("quarterly_data.xlsx")

# Pull in every sheet as a dictionary of DataFrames
sheets = pd.read_excel("quarterly_data.xlsx", sheet_name=None)

# Quick inspection
df.head()
df.info()
df.describe()

# Export after transformations
df.to_excel("cleaned_output.xlsx", index=False)
```

### Tips for robust reads

- Pin column types to avoid silent misinterpretation: `pd.read_excel("f.xlsx", dtype={"account_id": str})`
- Limit memory usage on wide files by selecting only needed columns: `usecols=["Date", "Amount", "Category"]`
- Parse date columns at read time: `parse_dates=["transaction_date"]`

---

## The Cardinal Rule: Formulas Belong in Excel, Not in Python

Whenever a cell's value depends on other cells, express that relationship as an Excel formula. Never compute the result in Python and paste the scalar into the workbook — doing so creates a static snapshot that breaks the moment upstream data changes.

### How this goes wrong

```python
# Avoid: computing in Python, writing a dead number
revenue_total = df["Revenue"].sum()
ws["F25"] = revenue_total  # writes 482000 — won't update if data changes

# Avoid: calculating a ratio in Python
margin = net_income / revenue
ws["G10"] = margin  # writes 0.12 — frozen forever
```

### How to do it right

```python
# Prefer: let the spreadsheet own the calculation
ws["F25"] = "=SUM(F3:F24)"

# Prefer: the ratio stays live
ws["G10"] = "=G8/G6"
```

This principle applies to every derived value — totals, averages, growth rates, percentage splits, running balances, variance calculations, all of it.

---

## End-to-End Workflow

1. **Pick your tool.** Data wrangling → pandas. Formatting and formulas → openpyxl. Often both.
2. **Open or create** the workbook.
3. **Populate and style** — write data, insert formulas, apply formatting.
4. **Save** the `.xlsx` file.
5. **Verify formulas** — import `formula_eval` and call `evaluate()` to compute every formula in memory and check for errors. Fix any issues and re-verify until the report comes back clean.

---

## Building a New Workbook

```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Header row
headers = ["Quarter", "Revenue", "COGS", "Gross Profit"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, name="Arial")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center")

# Sample data
ws.append(["Q1", 150000, 90000])
ws.append(["Q2", 175000, 100000])

# Gross profit formula for each row
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=4).value = f"=B{r}-C{r}"

# Totals
total_row = ws.max_row + 1
ws.cell(row=total_row, column=1, value="Total")
ws.cell(row=total_row, column=2).value = f"=SUM(B2:B{total_row - 1})"
ws.cell(row=total_row, column=3).value = f"=SUM(C2:C{total_row - 1})"
ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{total_row - 1})"

# Adjust column widths
for col_letter in ["A", "B", "C", "D"]:
    ws.column_dimensions[col_letter].width = 16

wb.save("quarterly_summary.xlsx")
```

---

## Modifying an Existing Workbook

```python
from openpyxl import load_workbook

wb = load_workbook("quarterly_summary.xlsx")
ws = wb["Summary"]

# Walk through every sheet
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"Processing: {name}, rows={sheet.max_row}")

# Update a value
ws["B2"] = 162000

# Structural changes
ws.insert_rows(3)          # push row 3 down, insert blank
ws.delete_cols(5)          # remove column E

# Add a new tab
notes = wb.create_sheet("Notes")
notes["A1"] = "Model last updated 2026-04-01"

wb.save("quarterly_summary_v2.xlsx")
```

---

## Verifying Formulas with `formula_eval`

openpyxl writes formula strings into cells but does **not** evaluate them. To verify that every formula resolves correctly, use the `formula_eval` module which computes all values in pure Python — no external applications needed.

```python
from formula_eval import evaluate

report = evaluate("quarterly_summary.xlsx")

if report["ok"]:
    print("All formulas clean")
else:
    for token, info in report["breakdown"].items():
        print(f"{token}: {info['count']} error(s)")
        for cell in info["cells"]:
            print(f"  {cell}")
```

You can also inspect individual computed values programmatically:

```python
report = evaluate("quarterly_summary.xlsx")

# Check a specific cell's computed result
assert report["computed"]["Summary!D4"] == 325000
```

### Understanding the report

The `evaluate()` function returns a dictionary with these keys:

| Key | Type | Description |
|---|---|---|
| `ok` | bool | `True` when every formula resolved without an error token |
| `formulas` | int | Total formula cells found in the workbook |
| `errors` | int | How many cells evaluated to an Excel error |
| `breakdown` | dict | Each error token mapped to `{"count": N, "cells": [...]}` |
| `computed` | dict | Every evaluated cell as `"Sheet!Cell": value` — useful for assertions |

A clean report:

```python
{
    "ok": True,
    "formulas": 18,
    "errors": 0,
    "breakdown": {},
    "computed": {"Summary!B4": 325000, "Summary!C4": 190000, ...}
}
```

A report with problems:

```python
{
    "ok": False,
    "formulas": 18,
    "errors": 3,
    "breakdown": {
        "#DIV/0!": {"count": 1, "cells": ["Summary!E14"]},
        "#REF!":   {"count": 2, "cells": ["Summary!D22", "Assumptions!B9"]}
    },
    "computed": {...}
}
```

When errors appear: fix the offending formulas in openpyxl, save again, and re-run `evaluate()` until `errors` reaches zero.

### Ensuring spreadsheet apps recalculate on open

Excel and Google Sheets automatically recalculate formulas when they open a file, but if you want to be explicit about it:

```python
from formula_eval import ensure_recalc_on_open

ensure_recalc_on_open("quarterly_summary.xlsx")
```

This sets the `fullCalcOnLoad` flag in the workbook metadata, guaranteeing that any spreadsheet application will do a fresh computation pass when the file is opened.

---

## Formula Debugging Checklist

### Before you build

- Spot-check two or three cell references to make sure they resolve to the values you expect.
- Confirm your column-number-to-letter mapping — column 64 is `BL`, not `BK`.
- Account for the index offset between pandas (0-based) and Excel (1-based). A DataFrame's row index 5 lands on Excel row 6.

### Common traps

- **Null values**: screen with `pd.notna()` before writing to cells.
- **Wide datasets**: fiscal-year data frequently sits beyond column 50 — don't stop scanning early.
- **Duplicate matches**: when searching for a label, verify you've found the right occurrence if it appears more than once.
- **Division by zero**: guard every divisor — wrap in `IFERROR` or check the denominator cell.
- **Broken references**: after inserting or deleting rows/columns, confirm existing formulas still point where they should.
- **Multi-sheet links**: the correct syntax is `SheetName!A1`; forgetting the sheet name prefix silently targets the active sheet.

### Incremental testing strategy

- Wire up formulas in a handful of cells first. Recalculate and verify before copying the pattern across hundreds of rows.
- Make sure every cell that a formula depends on actually exists and contains the expected type of data.
- Stress-test with boundary values: zero, negative numbers, very large figures.

---

## openpyxl Essentials

- Rows and columns are **1-indexed**: `cell(row=1, column=1)` is `A1`.
- To read previously computed values without formulas, open with `data_only=True`. **Caution**: saving a workbook opened this way permanently replaces formulas with their last-calculated values.
- For very large files, `read_only=True` (reading) and `write_only=True` (writing) dramatically reduce memory consumption.
- Formulas written by openpyxl are stored as text — use `formula_eval.evaluate()` to verify their computed values in Python.

## pandas Essentials

- Force column types at load time to prevent silent coercion: `dtype={"zip_code": str}`.
- Narrow wide files by selecting only relevant columns: `usecols=["A", "D", "F"]`.
- Let pandas handle date parsing: `parse_dates=["order_date"]`.

---

## Code Style Notes

When writing Python for spreadsheet tasks, favor brevity: short variable names, minimal inline commentary, no diagnostic print statements unless you're actively debugging. The code is a means to an end — the workbook is the deliverable.

Inside the workbook itself, be generous with documentation: annotate complex formulas with cell comments, tag every hardcoded input with its source, and add section labels so future readers can navigate the model without reverse-engineering it.
""",
    packages=["openpyxl", "pandas", "formulas"],
    modules={"formula_eval": str(_MODULES_DIR / "formula_eval.py")},
)
