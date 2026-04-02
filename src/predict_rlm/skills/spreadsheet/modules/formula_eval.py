"""Spreadsheet Formula Evaluation Engine

Pure-Python module for evaluating Excel formulas and verifying workbook
integrity. No external applications required — uses the `formulas` library
to build a dependency graph and compute every formula cell in memory.

Typical usage:

    from formula_eval import evaluate

    report = evaluate("model.xlsx")

    if report["ok"]:
        print("All formulas computed cleanly")
    else:
        for token, info in report["breakdown"].items():
            print(f"{token}: {info['count']} occurrence(s)")
            for loc in info["cells"]:
                print(f"  - {loc}")

    # Grab individual computed values
    print(report["computed"])  # {"Sheet!A3": 30.0, "Sheet!B1": 20.0, ...}

Dependencies: openpyxl, formulas (pip install formulas)
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

try:
    import formulas as _formulas_lib  # pyright: ignore[reportMissingImports]
except ImportError:
    raise ImportError(
        "The 'formulas' package is required. Install it with: pip install formulas"
    )

# Every standard Excel error token we scan for.
_ERROR_TOKENS = frozenset(["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"])

# Pattern for extracting sheet + cell from the formulas library key format:
#   '[workbook.xlsx]SheetName'!A1       -> (SheetName, A1)
#   '[workbook.xlsx]SheetName'!A1:A10   -> (SheetName, A1:A10)  (range — skip)
_CELL_KEY_RE = re.compile(r"'?\[.*?\](.+?)'?!([A-Z]+\d+)$")


def _extract_scalar(value: Any) -> Any:
    """Pull a plain Python scalar out of whatever the formulas lib returns."""
    if hasattr(value, "value"):
        value = value.value
    if isinstance(value, np.ndarray):
        value = value.flat[0]
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    return value


def _is_xl_error(value: Any) -> str | None:
    """If *value* is an Excel error, return the token string; else None."""
    s = str(value)
    return s if s in _ERROR_TOKENS else None


def _compute(path: str) -> dict[str, Any]:
    """Run the formulas engine and return a flat {key: scalar} mapping."""
    model = _formulas_lib.ExcelModel().loads(path).finish()
    raw = model.calculate()
    return {k: _extract_scalar(v) for k, v in raw.items()}


def _parse_cell_key(key: str) -> tuple[str, str] | None:
    """Extract (sheet_name, cell_ref) from a formulas-library result key.

    Returns None for range keys (e.g. A1:A10) since those are intermediate
    computations, not single-cell results.
    """
    m = _CELL_KEY_RE.match(key)
    if m:
        return m.group(1), m.group(2)
    return None


def _count_formulas(path: str) -> int:
    """Count cells whose value starts with '=' across every sheet."""
    wb = load_workbook(path, data_only=False)
    n = 0
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n += 1
    wb.close()
    return n


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def evaluate(workbook_path: str) -> dict:
    """Evaluate every formula in *workbook_path* and return a diagnostic report.

    The returned dictionary contains:

        ok        – bool, True when every formula resolved without error.
        formulas  – int, number of formula cells in the workbook.
        errors    – int, how many cells evaluated to an Excel error token.
        breakdown – dict mapping each error token that appeared to
                    {"count": int, "cells": ["Sheet!A1", ...]}.
        computed  – dict mapping "Sheet!Cell" to the evaluated value for
                    every formula cell. Useful for programmatic assertions.

    Raises FileNotFoundError if the workbook doesn't exist.
    """
    p = Path(workbook_path)
    if not p.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    resolved = str(p.resolve())

    raw_values = _compute(resolved)

    # Organise results into friendly structures
    computed: dict[str, Any] = {}
    error_cells: dict[str, list[str]] = {}

    for key, value in raw_values.items():
        parsed = _parse_cell_key(key)
        if parsed is None:
            continue
        sheet, cell_ref = parsed
        label = f"{sheet}!{cell_ref}"

        err = _is_xl_error(value)
        if err:
            error_cells.setdefault(err, []).append(label)
        else:
            computed[label] = value

    total_errors = sum(len(v) for v in error_cells.values())
    breakdown = {
        token: {"count": len(cells), "cells": cells} for token, cells in error_cells.items()
    }

    return {
        "ok": total_errors == 0,
        "formulas": _count_formulas(resolved),
        "errors": total_errors,
        "breakdown": breakdown,
        "computed": computed,
    }


def ensure_recalc_on_open(workbook_path: str) -> None:
    """Set the workbook flag that forces Excel / Google Sheets to recalculate
    every formula when the file is opened.

    This is a lightweight alternative to caching computed values — it simply
    marks the file so that any spreadsheet application will do a full
    recalculation pass on load. The flag is already True by default in new
    openpyxl workbooks, but call this after editing an existing file to be safe.
    """
    wb = load_workbook(workbook_path)
    wb.calculation.fullCalcOnLoad = True
    wb.save(workbook_path)
    wb.close()
