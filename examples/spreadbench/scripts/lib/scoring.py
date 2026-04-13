"""Cell-level comparison for scoring spreadsheet outputs against ground truth.

Ported from the spreadbench reference ``evaluate_rlm.py`` +
``optimize_signature.score_workbooks``. The primary entry point is
:func:`score_workbooks`, which returns a continuous ``(ratio, message)``
tuple so callers can use it for both pass/fail eval and GEPA's
fractional scoring.
"""

from __future__ import annotations

import datetime
import os
import re
from typing import Any

import openpyxl

_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)

_CELL_REF_RE = re.compile(r"^\$?[A-Za-z]{1,4}\$?[0-9]{1,7}$")
_CELL_RANGE_RE = re.compile(
    r"^\$?[A-Za-z]{1,4}\$?[0-9]{1,7}:\$?[A-Za-z]{1,4}\$?[0-9]{1,7}$"
)
_COL_RANGE_RE = re.compile(r"^\$?[A-Za-z]{1,4}:\$?[A-Za-z]{1,4}$")
_ROW_RANGE_RE = re.compile(r"^\$?[0-9]{1,7}:\$?[0-9]{1,7}$")


def _datetime_to_float(dt: datetime.datetime) -> float:
    delta = dt - _EXCEL_EPOCH
    return delta.days + delta.seconds / 86400.0


def transform_value(v: Any) -> Any:
    """Normalize a cell value for comparison.

    Numbers are rounded to 2 decimals, datetimes collapse to Excel's
    float-days representation, times drop the sub-second tail, and
    numeric strings get the numeric rounding treatment so text-formatted
    numbers still match their numeric counterparts.
    """
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if isinstance(v, datetime.time):
        return str(v)[:-3]
    if isinstance(v, datetime.datetime):
        return round(_datetime_to_float(v), 0)
    if isinstance(v, str):
        try:
            return round(float(v), 2)
        except ValueError:
            return v
    return v


def compare_cell_value(v1: Any, v2: Any) -> bool:
    """Return True when two cell values are considered equal after normalization."""
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) is not type(v2):
        return False
    return v1 == v2


def col_num2name(n: int) -> str:
    """Convert a 1-indexed column number to its Excel letter name."""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name: str) -> int:
    """Convert an Excel column letter name to a 1-indexed column number."""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    """Parse ``"A1:AB12"`` into ``((start_col, start_row), (end_col, end_row))``."""
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (
        (col_name2num(start_col), int(start_row)),
        (col_name2num(end_col), int(end_row)),
    )


def generate_cell_names(range_str: str, ws: Any = None) -> list[str]:
    """Expand a range like ``"A1:B5"`` into every covered cell name.

    Column-only ranges (e.g. ``"A:G"``) need a worksheet so we can bound
    them to ``ws.max_row``.
    """
    range_str = range_str.strip()
    if ":" not in range_str:
        return [range_str]

    start, end = (part.strip() for part in range_str.split(":", 1))
    start_has_digits = any(char.isdigit() for char in start)
    end_has_digits = any(char.isdigit() for char in end)

    if not start_has_digits and not end_has_digits:
        if ws is None:
            raise ValueError(
                f"Column-only range '{range_str}' requires worksheet to "
                f"determine row bounds"
            )
        range_str = f"{start}1:{end}{ws.max_row}"

    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]


def _looks_like_reference(token: str) -> bool:
    if not token:
        return False
    stripped = token.strip()
    if not stripped:
        return False
    if "!" in stripped:
        return True
    core = stripped.strip("'\"").replace("$", "").strip()
    return bool(
        _CELL_REF_RE.match(core)
        or _CELL_RANGE_RE.match(core)
        or _COL_RANGE_RE.match(core)
        or _ROW_RANGE_RE.match(core)
    )


def split_answer_position(answer_position: str) -> list[str]:
    """Split ``answer_position`` on commas, rejoining fragments that break
    inside sheet names with embedded commas."""
    raw_fragments = answer_position.split(",")
    parts: list[str] = []
    buffer: list[str] = []

    for fragment in raw_fragments:
        buffer.append(fragment)
        candidate = ",".join(buffer)
        if _looks_like_reference(candidate):
            cleaned = candidate.strip()
            if cleaned:
                parts.append(cleaned)
            buffer = []

    if buffer:
        dangling = ",".join(buffer).strip()
        if dangling:
            if parts:
                parts[-1] = f"{parts[-1]},{dangling}"
            else:
                parts.append(dangling)

    return parts


def _ensure_cell(cell_obj: Any) -> Any:
    """openpyxl returns tuples for range lookups; unwrap to a single cell."""
    while isinstance(cell_obj, tuple):
        if not cell_obj:
            return None
        cell_obj = cell_obj[0]
    return cell_obj


def score_workbooks(
    gt_file: str,
    proc_file: str,
    instruction_type: str,
    answer_position: str,
) -> tuple[float, str]:
    """Score *proc_file* against *gt_file* on the cells in *answer_position*.

    Returns ``(ratio, message)`` where ratio is ``matched_cells /
    total_cells`` across every ``sheet!range`` fragment in
    *answer_position*. A ratio of ``1.0`` means every compared cell
    matched — the "hard restriction" pass condition. ``instruction_type``
    is currently accepted for API parity with the reference and ignored
    here.
    """
    del instruction_type  # kept for caller parity; not used by this scorer

    if not os.path.exists(proc_file):
        return 0.0, "Output file does not exist"
    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return 0.0, f"Failed to load workbooks: {e}"

    total_cells = 0
    matched_cells = 0
    msgs: list[str] = []

    try:
        for part in split_answer_position(answer_position):
            part = part.strip()
            if "!" in part:
                sheet_name, cell_range = part.split("!", 1)
                sheet_name = sheet_name.strip().strip("'")
            else:
                sheet_name = wb_gt.sheetnames[0]
                cell_range = part
            cell_range = cell_range.strip().strip("'")

            if sheet_name not in wb_proc:
                cell_names = generate_cell_names(
                    cell_range, ws=wb_gt[sheet_name]
                )
                total_cells += len(cell_names)
                msgs.append(
                    f"Sheet '{sheet_name}' not found in output "
                    f"(has: {wb_proc.sheetnames})"
                )
                continue

            ws_gt = wb_gt[sheet_name]
            ws_proc = wb_proc[sheet_name]
            cell_names = generate_cell_names(cell_range, ws=ws_gt)
            diffs: list[str] = []

            for cell_name in cell_names:
                total_cells += 1
                cell_gt = _ensure_cell(ws_gt[cell_name])
                cell_proc = _ensure_cell(ws_proc[cell_name])
                if compare_cell_value(cell_gt.value, cell_proc.value):
                    matched_cells += 1
                else:
                    diffs.append(
                        f"  {sheet_name}!{cell_name}: "
                        f"expected={cell_gt.value!r} got={cell_proc.value!r}"
                    )

            if diffs:
                msgs.append(
                    f"Sheet '{sheet_name}' range {cell_range}: "
                    f"{matched_cells}/{total_cells} cells match\n"
                    + "\n".join(diffs[:20])
                )
                if len(diffs) > 20:
                    msgs.append(f"  ... ({len(diffs) - 20} more diffs)")
    finally:
        wb_gt.close()
        wb_proc.close()

    ratio = matched_cells / total_cells if total_cells > 0 else 0.0
    msg = "\n".join(msgs) if msgs else f"All {total_cells} cells match"
    return ratio, msg
