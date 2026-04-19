"""Eval loop for the SpreadsheetBench pipeline.

Call :func:`run_evaluation` with an :class:`EvalConfig` to run the
SpreadsheetRLM over a dataset, recalculate formulas on every produced
output via the local formulas+LO pipeline, and score against ground
truth. Works with either the seed docstring or a GEPA-optimized prompt
extracted from a run directory.
"""

from __future__ import annotations

import asyncio
import contextvars
import json
import logging
import os
import pickle
import shutil
import tempfile
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import dspy
import nest_asyncio
from spreadsheet_rlm.recalculate import recalculate
from spreadsheet_rlm.signature import ManipulateSpreadsheet
from spreadsheet_rlm.skills import libreoffice_spreadsheet_skill
from tqdm import tqdm

from predict_rlm import File, PredictRLM, Skill

from .dataset import SpreadsheetTask, load_dataset
from .lm_config import SUB_LM, compute_lm_cost, get_lm_config, get_sub_lm_config
from .scoring import score_workbooks

nest_asyncio.apply()

GEPA_COMPONENT_SIGNATURE = "signature_docstring"
GEPA_COMPONENT_SKILL = "skill_instructions"
# Backwards-compat alias used by older code that only knew about sig.
GEPA_COMPONENT = GEPA_COMPONENT_SIGNATURE

RLM_LOGGER_NAME = "dspy.predict.rlm"

_current_log_file: contextvars.ContextVar[Path | None] = contextvars.ContextVar(
    "spreadbench_eval_log_file", default=None
)


class _TaskFileHandler(logging.Handler):
    """Routes ``dspy.predict.rlm`` records to each async task's own log file.

    Every coroutine running concurrently under :func:`run_evaluation`
    lives in its own asyncio context, so the :class:`ContextVar` set by
    :func:`_run_case` is task-local. This handler reads that var on
    every emit and appends to the file the current task owns. When the
    var is ``None`` (logs disabled, or code running outside a task
    context), the record is dropped silently.
    """

    def emit(self, record: logging.LogRecord) -> None:
        log_file = _current_log_file.get()
        if log_file is None:
            return
        try:
            with open(log_file, "a") as f:
                f.write(self.format(record) + "\n")
        except Exception:
            self.handleError(record)


def _install_log_handler() -> _TaskFileHandler:
    """Attach the per-task file handler to ``dspy.predict.rlm``, idempotently.

    Clears any previously installed ``_TaskFileHandler`` first so that
    repeated :func:`run_evaluation` calls (from a notebook, say) don't
    stack handlers and duplicate records.
    """
    rlm_logger = logging.getLogger(RLM_LOGGER_NAME)
    rlm_logger.setLevel(logging.INFO)
    for existing in list(rlm_logger.handlers):
        if isinstance(existing, _TaskFileHandler):
            rlm_logger.removeHandler(existing)
    handler = _TaskFileHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    rlm_logger.addHandler(handler)
    return handler


@dataclass
class EvalConfig:
    """Knobs for a single eval run."""

    lm: str = "openai/gpt-5.4"
    sub_lm: str = SUB_LM
    reasoning_effort: str | None = "low"
    thinking_budget: int | None = None
    dataset: str = "testset"
    run_dir: str | None = None
    # When `run_dir` is set, by default both evolved components (signature
    # docstring + skill instructions) are loaded from the GEPA state and
    # applied. Pass ``only="signature"`` or ``only="skill"`` to apply just
    # one and use the seed value for the other — useful for A/B'ing each
    # component's individual contribution.
    only: str | None = None
    limit: int | None = None
    task_ids: tuple[str, ...] | None = None
    cases_per_task: int = 0
    """Cap on test cases per task (0 = all cases, matching the dataset loader default)."""
    concurrency: int = 30
    max_iterations: int = 50
    task_timeout: int = 300
    cache: bool = True
    log_dir: Path | None = None


@dataclass
class LMCost:
    """Aggregate cost + token usage for a single ``dspy.LM``."""

    role: str  # "main", "sub", "proposer", ...
    model: str
    calls: int
    prompt_tokens: int
    completion_tokens: int
    cost_usd: float

    @property
    def total_tokens(self) -> int:
        return self.prompt_tokens + self.completion_tokens

    def to_dict(self) -> dict:
        return {
            "role": self.role,
            "model": self.model,
            "calls": self.calls,
            "prompt_tokens": self.prompt_tokens,
            "completion_tokens": self.completion_tokens,
            "cost_usd": self.cost_usd,
        }


def summarize_lm_cost(role: str, lm: dspy.LM) -> LMCost:
    """Aggregate cost + tokens from a ``dspy.LM``'s call history.

    DSPy's ``LM.history`` records one dict per call (via LiteLLM), with
    ``cost`` and ``usage.prompt_tokens`` / ``usage.completion_tokens``.
    Models that LiteLLM doesn't have pricing for (e.g. mercury-2 via
    Inception Labs) report ``cost = None`` per call; for those models we
    fall back to :func:`compute_lm_cost`, which applies the per-model
    USD/MTok overrides declared in ``lm_config``.
    """
    history = getattr(lm, "history", []) or []
    calls = len(history)
    prompt_tokens = 0
    completion_tokens = 0
    cost_from_litellm = 0.0
    for entry in history:
        usage = entry.get("usage") or {}
        prompt_tokens += int(usage.get("prompt_tokens") or 0)
        completion_tokens += int(usage.get("completion_tokens") or 0)
        cost_from_litellm += float(entry.get("cost") or 0.0)

    model = getattr(lm, "model", "unknown")
    override_cost = compute_lm_cost(model, prompt_tokens, completion_tokens)
    cost = override_cost if override_cost is not None else cost_from_litellm

    return LMCost(
        role=role,
        model=model,
        calls=calls,
        prompt_tokens=prompt_tokens,
        completion_tokens=completion_tokens,
        cost_usd=cost,
    )


def _write_eval_trace_event(
    log_dir: Path,
    task_results: list["TaskResult"],
    total: int,
) -> None:
    """Append per-(role, model) JSONL rows summarising one eval run.

    Reads ``run_trace`` attached to each CaseResult by ``_run_case`` and
    sums the token/cost across all completed cases. Best-effort; any
    failure is swallowed so the eval's main path is unaffected.
    """
    try:
        from predict_rlm.trace import TokenUsage

        main_usage = TokenUsage()
        sub_usage = TokenUsage()
        main_model = ""
        sub_model: str | None = None
        traces_seen = 0
        for tr in task_results:
            for c in tr.cases:
                rt = getattr(c, "run_trace", None)
                if rt is None:
                    continue
                traces_seen += 1
                if not main_model:
                    main_model = str(getattr(rt, "model", ""))
                    sub_model = getattr(rt, "sub_model", None)
                main_usage += rt.usage.main
                if rt.usage.sub is not None:
                    sub_usage += rt.usage.sub

        if traces_seen == 0:
            return

        from datetime import datetime as _dt
        ts = _dt.now().isoformat()
        rows = []
        if main_usage.input_tokens or main_usage.output_tokens:
            rows.append({
                "ts": ts, "event": "evaluate", "role": "main", "model": main_model,
                "input_tokens": int(main_usage.input_tokens),
                "output_tokens": int(main_usage.output_tokens),
                "cost_usd": float(main_usage.cost),
                "tasks": total, "traces_captured": traces_seen,
            })
        if sub_model and (sub_usage.input_tokens or sub_usage.output_tokens):
            rows.append({
                "ts": ts, "event": "evaluate", "role": "sub", "model": sub_model,
                "input_tokens": int(sub_usage.input_tokens),
                "output_tokens": int(sub_usage.output_tokens),
                "cost_usd": float(sub_usage.cost),
                "tasks": total, "traces_captured": traces_seen,
            })

        log_path = log_dir / "cost_log.jsonl"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a") as f:
            for row in rows:
                f.write(json.dumps(row, default=str) + "\n")
    except Exception:
        pass  # best-effort observability


def _dump_eval_task_traces(log_dir: Path, task_results: list["TaskResult"]) -> None:
    """Write ``{log_dir}/task_traces.jsonl`` with one row per case.

    Each row carries the full RunTrace (serialized via
    ``to_exportable_json``) plus task_id / case_idx / score metadata so
    the log is self-describing. Best-effort; swallows errors.
    """
    try:
        out = log_dir / "task_traces.jsonl"
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w") as f:
            for tr in task_results:
                for c in tr.cases:
                    rt = getattr(c, "run_trace", None)
                    row: dict = {
                        "task_id": tr.task_id,
                        "case_idx": c.idx,
                        "score": c.score,
                        "passed": c.passed,
                        "message": c.message,
                        "recalc_source": c.recalc_source,
                    }
                    if rt is not None:
                        try:
                            row["trace"] = json.loads(
                                rt.to_exportable_json(indent=0)
                            )
                        except Exception:
                            row["trace"] = None
                    else:
                        row["trace"] = None
                    f.write(json.dumps(row, default=str) + "\n")
    except Exception:
        pass  # best-effort observability


@dataclass
class CaseResult:
    idx: int
    score: float
    passed: bool
    message: str
    recalc_source: str | None = None
    log_file: str | None = None


@dataclass
class TaskResult:
    task_id: str
    soft: float  # fraction of test cases that passed
    hard: int  # 1 iff every test case passed
    cases: list[CaseResult] = field(default_factory=list)


@dataclass
class EvalReport:
    config: EvalConfig
    signature_source: str
    signature_length: int
    skill_source: str
    skill_length: int
    total_tasks: int
    soft_restriction_avg: float
    hard_restriction_avg: float
    tasks_all_passing: int
    duration_seconds: float
    per_task: list[TaskResult]
    costs: list[LMCost]

    @property
    def total_cost_usd(self) -> float:
        return sum(c.cost_usd for c in self.costs)

    def to_dict(self) -> dict:
        return {
            "config": {
                "lm": self.config.lm,
                "sub_lm": self.config.sub_lm,
                "reasoning_effort": self.config.reasoning_effort,
                "dataset": self.config.dataset,
                "run_dir": self.config.run_dir,
                "only": self.config.only,
                "limit": self.config.limit,
                "concurrency": self.config.concurrency,
                "max_iterations": self.config.max_iterations,
                "task_timeout": self.config.task_timeout,
            },
            "signature_source": self.signature_source,
            "signature_length": self.signature_length,
            "skill_source": self.skill_source,
            "skill_length": self.skill_length,
            "total_tasks": self.total_tasks,
            "soft_restriction_avg": self.soft_restriction_avg,
            "hard_restriction_avg": self.hard_restriction_avg,
            "tasks_all_passing": self.tasks_all_passing,
            "duration_seconds": self.duration_seconds,
            "costs": [c.to_dict() for c in self.costs],
            "total_cost_usd": self.total_cost_usd,
            "per_task": [
                {
                    "task_id": t.task_id,
                    "soft": t.soft,
                    "hard": t.hard,
                    "cases": [
                        {
                            "idx": c.idx,
                            "score": c.score,
                            "passed": c.passed,
                            "message": c.message,
                            "recalc_source": c.recalc_source,
                            "log_file": c.log_file,
                        }
                        for c in t.cases
                    ],
                }
                for t in self.per_task
            ],
        }


def make_dynamic_signature(docstring: str) -> type[dspy.Signature]:
    """Return a ``ManipulateSpreadsheet`` variant with a custom docstring."""
    return ManipulateSpreadsheet.with_instructions(docstring)


def _build_instruction(
    instruction: str,
    answer_range: str,
    answer_sheet: str,
    instruction_type: str,
) -> str:
    if instruction_type == "Sheet-Level Manipulation":
        pos_note = (
            f"Answer position: {answer_range} on sheet '{answer_sheet}'. "
            "This is the maximum range of cells you may modify. "
            "You only need to modify or fill in values within this range."
        )
    else:
        pos_note = (
            f"Answer position: {answer_range} on sheet '{answer_sheet}'. "
            "This is the cell position to be modified or filled. "
            "You only need to modify or fill in values within this range."
        )
    return f"{instruction}\n\n{pos_note}"


def parse_answer_position(
    answer_position: str, input_path: str
) -> tuple[str, str]:
    """Split *answer_position* into ``(sheet_name, cell_range)``.

    Mirrors the grader's parse logic in ``lib.scoring``:

    * ``"'output'!A2:G15"`` → ``("output", "A2:G15")``
    * ``"Sheet1'!A1:F14"``  → ``("Sheet1", "A1:F14")`` (handles a malformed
      closing-only quote seen in ~19 tasks)
    * ``"A3:D32"``          → first sheet of the input workbook + original
      range
    * ``"'OUT CAS'!A2:C10,'OUT CAS'!E2:G5"`` → first fragment's sheet, plus
      the original multi-range string so the model sees every range

    Falls back to ``"Sheet1"`` if the input workbook cannot be opened.
    """
    from openpyxl import load_workbook

    has_comma = "," in answer_position
    first_fragment = answer_position.split(",", 1)[0].strip()

    if "!" in first_fragment:
        sheet_part, range_part = first_fragment.split("!", 1)
        sheet_name = sheet_part.strip().strip("'")
        if has_comma:
            return sheet_name, answer_position
        return sheet_name, range_part.strip()

    try:
        wb = load_workbook(input_path, read_only=True)
        try:
            sheet_name = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
        finally:
            wb.close()
    except Exception:
        sheet_name = "Sheet1"
    return sheet_name, answer_position


def make_dynamic_skill(instructions: str) -> Skill:
    """Return a copy of the LO spreadsheet skill with custom instructions.

    Preserves ``name``, ``packages``, ``modules``, and ``tools`` (including
    the ``recalculate`` host-side tool registered by ``skills.py``). Uses
    pydantic ``model_copy`` so the new Skill is structurally identical to
    the seed except for the ``instructions`` field.
    """
    return libreoffice_spreadsheet_skill.model_copy(
        update={"instructions": instructions}
    )


def extract_best_candidate(run_dir: str | Path) -> dict[str, str]:
    """Pull the best-by-mean candidate (full dict) from a GEPA ``run_dir``.

    Returns the full ``{component_name: component_text}`` dict. For
    multi-component runs this includes both ``signature_docstring`` and
    ``skill_instructions``; for single-component (legacy) runs only
    ``signature_docstring`` is present.
    """
    state_path = Path(run_dir) / "gepa_state.bin"
    with state_path.open("rb") as f:
        state = pickle.load(f)
    subs = state["prog_candidate_val_subscores"]
    cands = state["program_candidates"]
    best_idx = max(
        range(len(cands)),
        key=lambda i: (sum(subs[i].values()) / len(subs[i])) if subs[i] else 0.0,
    )
    return cands[best_idx]


def extract_best_prompt(run_dir: str | Path) -> tuple[str, int, float]:
    """Pull the best-by-mean candidate prompt from a GEPA ``run_dir``.

    Reads ``{run_dir}/gepa_state.bin`` (a pickled GEPA state dict),
    picks the candidate whose mean validation subscore is highest,
    and returns ``(prompt_text, candidate_index, mean_score)``.
    """
    state_path = Path(run_dir) / "gepa_state.bin"
    with state_path.open("rb") as f:
        state = pickle.load(f)
    subs = state["prog_candidate_val_subscores"]
    cands = state["program_candidates"]
    best_idx = max(
        range(len(cands)),
        key=lambda i: (sum(subs[i].values()) / len(subs[i])) if subs[i] else 0.0,
    )
    best_mean = sum(subs[best_idx].values()) / len(subs[best_idx])
    return cands[best_idx][GEPA_COMPONENT], best_idx, best_mean


async def _run_case(
    task: SpreadsheetTask,
    idx: int,
    input_path: str,
    answer_path: str | None,
    sig_cls: type[dspy.Signature],
    skill: Skill,
    lm: dspy.LM,
    sub_lm: dspy.LM,
    sem: asyncio.Semaphore,
    tmp_dir: str,
    config: EvalConfig,
) -> CaseResult:
    output_path = os.path.join(tmp_dir, f"{idx}_{task.task_id}_output.xlsx")

    log_file: Path | None = None
    log_file_str: str | None = None
    if config.log_dir is not None:
        log_file = config.log_dir / task.task_id / f"case_{idx}.log"
        log_file.parent.mkdir(parents=True, exist_ok=True)
        log_file.write_text(
            f"task: {task.task_id}  case: {idx}\n"
            f"instruction_type: {task.instruction_type}\n"
            f"answer_position: {task.answer_position}\n"
            f"input:  {input_path}\n"
            f"answer: {answer_path}\n"
            f"{'=' * 60}\n"
        )
        _current_log_file.set(log_file)
        log_file_str = str(log_file)

    answer_sheet, answer_range = await asyncio.to_thread(
        parse_answer_position, task.answer_position, input_path
    )
    formatted_instruction = _build_instruction(
        task.instruction, answer_range, answer_sheet, task.instruction_type
    )

    run_trace: Any = None
    async with sem:
        try:
            predictor = PredictRLM(
                sig_cls,
                lm=lm,
                sub_lm=sub_lm,
                skills=[skill],
                max_iterations=config.max_iterations,
                verbose=config.log_dir is not None,
                debug=False,
            )
            result = await asyncio.wait_for(
                predictor.acall(
                    input_spreadsheet=File(path=input_path),
                    instruction=formatted_instruction,
                ),
                timeout=config.task_timeout,
            )
            run_trace = getattr(result, "trace", None)
            if not (
                result
                and result.output_spreadsheet
                and result.output_spreadsheet.path
                and os.path.exists(result.output_spreadsheet.path)
            ):
                cr = CaseResult(idx, 0.0, False, "No output", log_file=log_file_str)
                cr.run_trace = run_trace  # type: ignore[attr-defined]
                return cr
            shutil.copy2(result.output_spreadsheet.path, output_path)
        except asyncio.TimeoutError as e:
            from predict_rlm.trace import extract_trace_from_exc

            cr = CaseResult(
                idx, 0.0, False, f"Timeout ({config.task_timeout}s)",
                log_file=log_file_str,
            )
            cr.run_trace = extract_trace_from_exc(e)  # type: ignore[attr-defined]
            return cr
        except Exception as e:
            from predict_rlm.trace import extract_trace_from_exc

            cr = CaseResult(
                idx, 0.0, False, f"RLM error: {e}", log_file=log_file_str,
            )
            cr.run_trace = extract_trace_from_exc(e)  # type: ignore[attr-defined]
            return cr

    recalc_source: str | None = None
    try:
        recalc_result = await asyncio.to_thread(recalculate, output_path)
        recalc_source = recalc_result.source
    except Exception as e:
        recalc_source = f"failed: {e}"

    if answer_path is None:
        cr = CaseResult(
            idx, 0.0, False, "Answer file not found",
            recalc_source=recalc_source, log_file=log_file_str,
        )
        cr.run_trace = run_trace  # type: ignore[attr-defined]
        return cr

    try:
        ratio, msg = await asyncio.to_thread(
            score_workbooks,
            answer_path,
            output_path,
            task.instruction_type,
            task.answer_position,
        )
        if log_file is not None:
            with log_file.open("a") as f:
                f.write(
                    f"\n{'=' * 60}\n"
                    f"recalc_source: {recalc_source}\n"
                    f"score:         {ratio:.4f}  "
                    f"{'PASS' if ratio == 1.0 else 'FAIL'}\n"
                    f"{msg}\n"
                )
        cr = CaseResult(
            idx, ratio, ratio == 1.0, msg,
            recalc_source=recalc_source, log_file=log_file_str,
        )
        cr.run_trace = run_trace  # type: ignore[attr-defined]
        return cr
    except Exception as e:
        cr = CaseResult(
            idx, 0.0, False, f"Comparison error: {e}",
            recalc_source=recalc_source, log_file=log_file_str,
        )
        cr.run_trace = run_trace  # type: ignore[attr-defined]
        return cr


async def _run_tasks_async(
    tasks: list[SpreadsheetTask],
    sig_cls: type[dspy.Signature],
    skill: Skill,
    lm: dspy.LM,
    sub_lm: dspy.LM,
    config: EvalConfig,
) -> list[TaskResult]:
    tmp_dir = tempfile.mkdtemp(prefix="spreadbench_eval_")
    sem = asyncio.Semaphore(config.concurrency)
    pbar = tqdm(total=len(tasks), desc="Evaluating", unit="task")

    async def _process(task: SpreadsheetTask) -> TaskResult:
        case_coros = [
            _run_case(
                task, idx, input_path, answer_path,
                sig_cls, skill, lm, sub_lm, sem, tmp_dir, config,
            )
            for idx, input_path, answer_path in task.test_cases
        ]
        cases = await asyncio.gather(*case_coros)
        soft = (
            sum(c.score for c in cases) / len(cases) if cases else 0.0
        )
        hard = 1 if cases and all(c.passed for c in cases) else 0
        pbar.set_postfix_str(f"{task.task_id}={soft:.0%}")
        pbar.update(1)
        return TaskResult(task_id=task.task_id, soft=soft, hard=hard, cases=list(cases))

    try:
        results = await asyncio.gather(*(_process(t) for t in tasks))
    finally:
        pbar.close()
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return list(results)


def _resolve_components(
    config: EvalConfig,
) -> tuple[str, str, str, str]:
    """Resolve (sig_text, sig_source, skill_text, skill_source) for an eval run.

    Default: load both components from ``config.run_dir`` if set, fall
    back to seed for either component that the run_dir doesn't contain
    (so single-component / legacy runs still work).

    ``config.only`` overrides this:
    * ``"signature"`` → load sig from run_dir, use seed skill
    * ``"skill"``     → use seed sig, load skill from run_dir
    * ``None``        → load both (default)

    Raises:
        ValueError: if ``only`` is set but ``run_dir`` is not, or if
            ``only`` is an unrecognised value.
    """
    seed_sig = ManipulateSpreadsheet.__doc__ or ""
    seed_skill = libreoffice_spreadsheet_skill.instructions

    if config.only is not None and config.only not in ("signature", "skill"):
        raise ValueError(
            f"--only must be 'signature' or 'skill', got {config.only!r}"
        )
    if config.only is not None and not config.run_dir:
        raise ValueError("--only requires --run_dir")

    if not config.run_dir:
        return seed_sig, "seed", seed_skill, "seed"

    candidate = extract_best_candidate(config.run_dir)
    run_basename = Path(config.run_dir).name

    sig_text = seed_sig
    sig_source = "seed"
    if config.only != "skill" and GEPA_COMPONENT_SIGNATURE in candidate:
        sig_text = candidate[GEPA_COMPONENT_SIGNATURE]
        sig_source = f"{run_basename}#sig"

    skill_text = seed_skill
    skill_source = "seed"
    if config.only != "signature" and GEPA_COMPONENT_SKILL in candidate:
        skill_text = candidate[GEPA_COMPONENT_SKILL]
        skill_source = f"{run_basename}#skill"

    return sig_text, sig_source, skill_text, skill_source


def run_evaluation(config: EvalConfig) -> EvalReport:
    """Run the eval loop described by *config* and return a report.

    Resolves both components (signature docstring + skill instructions)
    from ``config.run_dir`` when set, builds the main/sub LMs, loads the
    dataset, runs every task's cases concurrently, recalculates each
    produced output through the formulas+LO pipeline, and scores
    against ground truth. ``config.only`` selectively applies just one
    of the two evolved components.
    """
    sig_text, sig_source, skill_text, skill_source = _resolve_components(config)
    print(
        f"Signature: {sig_source} ({len(sig_text)} chars)\n"
        f"Skill:     {skill_source} ({len(skill_text)} chars)"
    )

    sig_cls = make_dynamic_signature(sig_text)
    skill = make_dynamic_skill(skill_text)

    lm_cfg = get_lm_config(
        config.lm,
        config.reasoning_effort,
        thinking_budget=config.thinking_budget,
    )
    sub_lm_cfg = get_sub_lm_config(config.sub_lm)
    lm = dspy.LM(**lm_cfg, cache=config.cache)
    sub_lm = dspy.LM(**sub_lm_cfg, cache=config.cache)

    effort = config.reasoning_effort if config.reasoning_effort else "none"
    print(f"Main LM: {config.lm}  (reasoning_effort={effort})")
    print(f"Sub LM:  {config.sub_lm}")

    tasks = load_dataset(config.dataset, max_cases_per_task=config.cases_per_task)
    if config.task_ids:
        wanted = set(config.task_ids)
        tasks = [t for t in tasks if t.task_id in wanted]
        missing = wanted - {t.task_id for t in tasks}
        if missing:
            raise ValueError(
                f"task_ids not found in dataset {config.dataset!r}: "
                f"{sorted(missing)}"
            )
    if config.limit:
        tasks = tasks[: config.limit]
    print(f"Dataset: {config.dataset} ({len(tasks)} tasks)")

    if config.log_dir is not None:
        config.log_dir.mkdir(parents=True, exist_ok=True)
        _install_log_handler()
        print(f"Per-case RLM logs: {config.log_dir}/<task_id>/case_<idx>.log")

    print(
        f"Running {len(tasks)} tasks with concurrency={config.concurrency}, "
        f"max_iterations={config.max_iterations}, "
        f"task_timeout={config.task_timeout}s..."
    )
    t0 = time.time()
    loop = asyncio.get_event_loop()
    results = loop.run_until_complete(
        _run_tasks_async(tasks, sig_cls, skill, lm, sub_lm, config)
    )
    elapsed = time.time() - t0

    total = len(results)
    soft_avg = sum(r.soft for r in results) / total if total else 0.0
    hard_avg = sum(r.hard for r in results) / total if total else 0.0
    hard_count = sum(r.hard for r in results)

    costs = [
        summarize_lm_cost("main", lm),
        summarize_lm_cost("sub", sub_lm),
    ]

    # Append one JSONL event row per (role, model) to config.log_dir/cost_log.jsonl
    # and dump full per-case RunTraces to config.log_dir/task_traces.jsonl. We
    # pull usage straight from each PredictRLM's trace rather than re-aggregating
    # from LM history.
    if config.log_dir is not None:
        _write_eval_trace_event(config.log_dir, results, total)
        _dump_eval_task_traces(config.log_dir, results)

    return EvalReport(
        config=config,
        signature_source=sig_source,
        signature_length=len(sig_text),
        skill_source=skill_source,
        skill_length=len(skill_text),
        total_tasks=total,
        soft_restriction_avg=soft_avg,
        hard_restriction_avg=hard_avg,
        tasks_all_passing=hard_count,
        duration_seconds=elapsed,
        per_task=results,
        costs=costs,
    )
