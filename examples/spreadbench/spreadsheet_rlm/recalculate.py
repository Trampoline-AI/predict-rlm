"""Formula recalculation pipeline for xlsx files.

Resolves stale formula values inside a workbook by running it through a
two-stage pipeline:

1. **Python `formulas` library** — computes every formula in memory
   using a pure-Python dependency graph. Fast, no subprocess, and
   handles a surprisingly wide range of Excel functions including
   `TEXTJOIN` and other modern built-ins.
2. **LibreOffice headless** — ``soffice --headless --convert-to xlsx``
   as a fallback for formulas the Python library can't evaluate.

The pipeline picks whichever candidate resolved the most formula cells,
with the untouched baseline always in the running so the call is
strictly additive: the winning file has at least as many resolved
formula cells as the input. This mirrors the "zero regressions" property
the spreadbench eval relies on — a recalc never turns a passing test
into a failing one by destroying already-cached values.

Typical usage::

    from spreadsheet_rlm.recalculate import recalculate

    result = recalculate("model_output.xlsx")
    # result.source      -> "baseline" | "formulas" | "libreoffice"
    # result.resolved    -> number of resolved formula cells in the winner
    # result.errors      -> list of per-stage failure messages (if any)
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import subprocess
import tempfile
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import openpyxl

try:
    import formulas as _formulas_lib
except ImportError:
    _formulas_lib = None

log = logging.getLogger("spreadsheet_rlm.recalculate")

_FORMULAS_TQDM_PATCHED = False


class _QuietTqdm:
    """No-op stand-in for ``tqdm.tqdm`` used by the formulas library.

    Mirrors the QuietTqdm pattern from ``formulas/cli.py`` — when monkey
    patched into the tqdm reference inside ``formulas.excel``, every
    progress bar call becomes a no-op. Used to keep the formulas
    library's internal progress bars from interleaving with the eval
    loop's own tqdm output during recalculation.
    """

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        pass

    def __enter__(self) -> "_QuietTqdm":
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> bool:
        return False

    def update(self, *args: Any, **kwargs: Any) -> None:
        return None

    def refresh(self) -> None:
        return None

    total = 0


def _silence_formulas_tqdm() -> None:
    global _FORMULAS_TQDM_PATCHED
    if _FORMULAS_TQDM_PATCHED or _formulas_lib is None:
        return
    _formulas_lib.ExcelModel.complete.__globals__["tqdm"].tqdm = _QuietTqdm
    _FORMULAS_TQDM_PATCHED = True

_ERROR_TOKENS = frozenset(
    ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
)

# Keys returned by the formulas library look like:
#   "'[workbook.xlsx]SHEETNAME'!A1"
# The sheet name is always upper-cased by the library.
_CELL_KEY_RE = re.compile(r"'?\[.*?\](.+?)'?!([A-Z]+\d+)$")

Source = Literal["baseline", "formulas", "libreoffice"]


@dataclass
class RecalcResult:
    """Outcome of a :func:`recalculate` call."""

    source: Source
    resolved: int
    total_formulas: int
    baseline_resolved: int
    formulas_resolved: int | None = None
    libreoffice_resolved: int | None = None
    errors: list[str] = field(default_factory=list)


def recalculate(path: str | Path) -> RecalcResult:
    """Recalculate formulas in *path*, writing the best candidate back in place.

    The pipeline snapshots the baseline workbook, runs the Python
    `formulas` library, and falls back to LibreOffice headless if any
    formulas were left unresolved. The candidate with the most resolved
    formula cells wins; baseline is always a candidate so the file is
    never downgraded.

    Returns a :class:`RecalcResult`. Ties prefer the least destructive
    source in the order ``baseline > libreoffice > formulas`` —
    baseline changes nothing, LibreOffice keeps formula strings intact,
    and the formulas library flattens computed formulas to literal
    values.
    """
    src = Path(path).resolve()
    if not src.is_file():
        raise FileNotFoundError(f"workbook not found: {src}")

    targets = _formula_targets(src)
    total = len(targets)

    if total == 0:
        return RecalcResult(
            source="baseline",
            resolved=0,
            total_formulas=0,
            baseline_resolved=0,
        )

    baseline_resolved = _count_resolved(src, targets)
    log.debug("baseline: %d/%d formula cells resolved", baseline_resolved, total)

    if baseline_resolved == total:
        return RecalcResult(
            source="baseline",
            resolved=baseline_resolved,
            total_formulas=total,
            baseline_resolved=baseline_resolved,
        )

    errors: list[str] = []
    candidates: dict[Source, tuple[Path, int]] = {}
    formulas_resolved: int | None = None
    libreoffice_resolved: int | None = None

    with tempfile.TemporaryDirectory() as tmpname:
        tmp = Path(tmpname)

        baseline_copy = tmp / "baseline.xlsx"
        shutil.copy2(src, baseline_copy)
        candidates["baseline"] = (baseline_copy, baseline_resolved)

        formulas_out = tmp / "formulas.xlsx"
        try:
            _recalc_with_formulas(src, formulas_out)
            formulas_resolved = _count_resolved(formulas_out, targets)
            candidates["formulas"] = (formulas_out, formulas_resolved)
            log.debug("formulas: %d/%d resolved", formulas_resolved, total)
        except Exception as e:
            errors.append(f"formulas: {e}")
            log.debug("formulas pipeline failed: %s", e)

        # Fast path: if the Python library resolved everything, skip the
        # LibreOffice subprocess entirely — that's the whole point of
        # running formulas first.
        if formulas_resolved == total:
            _replace(candidates["formulas"][0], src)
            return RecalcResult(
                source="formulas",
                resolved=formulas_resolved,
                total_formulas=total,
                baseline_resolved=baseline_resolved,
                formulas_resolved=formulas_resolved,
                libreoffice_resolved=None,
                errors=errors,
            )

        lo_out = tmp / "libreoffice.xlsx"
        try:
            _recalc_with_libreoffice(src, lo_out)
            libreoffice_resolved = _count_resolved(lo_out, targets)
            candidates["libreoffice"] = (lo_out, libreoffice_resolved)
            log.debug("libreoffice: %d/%d resolved", libreoffice_resolved, total)
        except Exception as e:
            errors.append(f"libreoffice: {e}")
            log.debug("libreoffice pipeline failed: %s", e)

        winner = _pick_winner(candidates)
        winner_path, winner_count = candidates[winner]
        if winner != "baseline":
            _replace(winner_path, src)

    return RecalcResult(
        source=winner,
        resolved=winner_count,
        total_formulas=total,
        baseline_resolved=baseline_resolved,
        formulas_resolved=formulas_resolved,
        libreoffice_resolved=libreoffice_resolved,
        errors=errors,
    )


def _formula_targets(path: Path) -> list[tuple[str, str]]:
    """Return ``[(sheet_name, cell_coordinate)]`` for every formula cell."""
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    result: list[tuple[str, str]] = []
    try:
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for cell in row:
                    val = cell.value
                    if not (isinstance(val, str) and val.startswith("=")):
                        continue
                    # In read-only mode empty cells are EmptyCell instances
                    # that don't carry a coordinate attribute; skip defensively.
                    coord = getattr(cell, "coordinate", None)
                    if coord is not None:
                        result.append((name, coord))
    finally:
        wb.close()
    return result


def _count_resolved(path: Path, targets: list[tuple[str, str]]) -> int:
    """Count *targets* whose cached value in *path* is a non-error, non-null."""
    if not targets:
        return 0

    wb = openpyxl.load_workbook(path, data_only=True)
    resolved = 0
    try:
        for sheet_name, coord in targets:
            if sheet_name not in wb.sheetnames:
                continue
            val = wb[sheet_name][coord].value
            if val is None:
                continue
            if isinstance(val, str) and val in _ERROR_TOKENS:
                continue
            resolved += 1
    finally:
        wb.close()
    return resolved


def _recalc_with_formulas(src: Path, dst: Path) -> None:
    """Run the Python `formulas` library on *src* and write the result to *dst*.

    Non-formula cells, sheet names, and styling are preserved by loading
    *src* with openpyxl and only replacing formula cells whose value the
    library successfully computed. Formulas the library couldn't handle
    are left untouched so the LibreOffice fallback still has something
    to work on.
    """
    if _formulas_lib is None:
        raise RuntimeError("formulas library not installed")

    _silence_formulas_tqdm()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        model = _formulas_lib.ExcelModel().loads(str(src)).finish()
        solution = model.calculate()

    computed: dict[tuple[str, str], Any] = {}
    for key, value in solution.items():
        m = _CELL_KEY_RE.match(key)
        if not m:
            continue
        sheet_upper, cell_ref = m.group(1), m.group(2)
        scalar = _extract_scalar(value)
        if scalar is None:
            continue
        if isinstance(scalar, str) and scalar in _ERROR_TOKENS:
            continue
        computed[(sheet_upper, cell_ref)] = scalar

    wb = openpyxl.load_workbook(src)
    try:
        for sheet in wb.worksheets:
            key_sheet = sheet.title.upper()
            for row in sheet.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    new_value = computed.get((key_sheet, cell.coordinate))
                    if new_value is not None:
                        cell.value = new_value
        wb.save(dst)
    finally:
        wb.close()


def _recalc_with_libreoffice(src: Path, dst: Path) -> None:
    """Recalculate *src* via ``soffice --headless`` and write it to *dst*."""
    soffice = _find_libreoffice()
    if not soffice:
        raise RuntimeError("LibreOffice not found")

    original_names: list[str] = []
    try:
        wb = openpyxl.load_workbook(src, read_only=True)
        original_names = list(wb.sheetnames)
        wb.close()
    except Exception:
        pass

    with tempfile.TemporaryDirectory() as tmpname:
        tmp = Path(tmpname)
        profile = tmp / "profile"
        profile.mkdir()
        outdir = tmp / "out"
        outdir.mkdir()
        result = subprocess.run(
            [
                soffice,
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx:Calc MS Excel 2007 XML",
                "--outdir",
                str(outdir),
                f"-env:UserInstallation=file://{profile}",
                str(src),
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            msg = result.stderr.strip() or result.stdout.strip() or "unknown error"
            raise RuntimeError(f"libreoffice failed: {msg}")
        produced = list(outdir.glob("*.xlsx"))
        if not produced:
            raise RuntimeError("libreoffice produced no xlsx output")
        shutil.copy2(produced[0], dst)

    if original_names:
        _restore_sheet_names(dst, original_names)


def _restore_sheet_names(path: Path, names: list[str]) -> None:
    """Rename sheets back to *names* if LibreOffice renamed them in place."""
    try:
        wb = openpyxl.load_workbook(path)
        if wb.sheetnames != names and len(wb.sheetnames) == len(names):
            for ws, orig in zip(wb.worksheets, names):
                if ws.title != orig:
                    ws.title = orig
            wb.save(path)
        wb.close()
    except Exception:
        pass


def _find_libreoffice() -> str | None:
    """Locate the LibreOffice binary, or return None if absent."""
    candidates = [
        "soffice",
        "libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
    ]
    for c in candidates:
        path = shutil.which(c)
        if path:
            return path
        if os.path.isfile(c):
            return c
    return None


def _pick_winner(candidates: dict[Source, tuple[Path, int]]) -> Source:
    """Pick the candidate with the most resolved cells.

    Iteration order ``baseline -> libreoffice -> formulas`` combined with a
    strict ``>`` comparison means ties are resolved in favour of the
    earlier (less destructive) candidate: baseline wins ties against
    everything, LibreOffice wins ties against formulas.
    """
    preference: list[Source] = ["baseline", "libreoffice", "formulas"]
    winner: Source = "baseline"
    winner_count = candidates["baseline"][1]
    for name in preference[1:]:
        if name not in candidates:
            continue
        count = candidates[name][1]
        if count > winner_count:
            winner = name
            winner_count = count
    return winner


def _extract_scalar(value: Any) -> Any:
    """Coerce a `formulas` library return value to a plain Python scalar."""
    try:
        import numpy as np
    except ImportError:
        np = None

    if hasattr(value, "value"):
        value = value.value
    if np is not None and isinstance(value, np.ndarray):
        if value.size == 0:
            return None
        value = value.flat[0]
    if np is not None:
        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.floating):
            return float(value)
    return value


def _replace(src: Path, dst: Path) -> None:
    """Overwrite *dst* with the contents of *src*, preserving metadata."""
    shutil.copy2(src, dst)
