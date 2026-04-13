"""Custom spreadsheet skill targeting LibreOffice compatibility."""

import os
from pathlib import Path
from typing import Annotated

from predict_rlm import Skill, SyncedFile

from .recalculate import recalculate as _recalc_impl

# Reuse the formula_eval module from the built-in spreadsheet skill
_BUILTIN_MODULES_DIR = (
    Path(__file__).parent.parent
    / ".venv"
    / "lib"
    / "python3.13"
    / "site-packages"
    / "predict_rlm"
    / "skills"
    / "spreadsheet"
    / "modules"
)

# Fallback: find it relative to the installed package
if not _BUILTIN_MODULES_DIR.exists():
    import importlib

    _builtin_skill_mod = importlib.import_module("predict_rlm.skills.spreadsheet.skill")

    _BUILTIN_MODULES_DIR = Path(_builtin_skill_mod.__file__).parent / "modules"


def recalculate(file_path: Annotated[Path, SyncedFile()]) -> str:
    """Recalculate all formulas in an xlsx file and cache the results in place.

    Runs a two-stage pipeline: the Python `formulas` library first (fast,
    in-process), then LibreOffice headless as a fallback for any formulas
    the library couldn't evaluate. The winning candidate is whichever
    resolved the most formula cells — the untouched original is always
    a candidate too, so the call is strictly additive and never downgrades
    a file. After this call, re-open the file with
    `load_workbook(file_path, data_only=True)` to read the cached values.

    Args:
        file_path: Path to the .xlsx file to recalculate.

    Returns:
        "ok" on success (optionally annotated with the source that won and
        the resolved/total counts), or an error message starting with
        "Error:".
    """
    path = os.path.abspath(file_path)
    if not os.path.isfile(path):
        return f"Error: file not found: {path}"
    try:
        result = _recalc_impl(path)
    except Exception as e:
        return f"Error: {e}"

    status = (
        f"ok (source={result.source}, "
        f"resolved={result.resolved}/{result.total_formulas})"
    )
    if result.errors:
        status += f" [warnings: {'; '.join(result.errors)}]"
    return status


libreoffice_spreadsheet_skill = Skill(
    name="spreadsheet-libreoffice",
    instructions="""Use openpyxl and pandas to build, transform, and style spreadsheet files (.xlsx).

Output files will be evaluated by opening them in **LibreOffice Calc** and reading the cached cell values. This has critical implications for formulas.

# Formula Compatibility — CRITICAL

## Spill Functions Do NOT Work

LibreOffice supports FILTER, SORT, UNIQUE, etc. as formulas, but **spill results are not persisted** when the file is saved as xlsx. Only the first cell of a spill array gets cached — all other spilled cells read as empty/None. This means any formula that returns an array into multiple cells **will silently lose data**.

**NEVER use these spill/dynamic-array functions in formulas:**
`FILTER`, `SORT`, `UNIQUE`, `SORTBY`, `SEQUENCE`, `RANDARRAY`

**Write every cell individually.** If a formula should fill a range (e.g., B3:B14), you must write the formula to every cell in that range in a loop. Writing only to the anchor cell leaves the rest empty:
```python
# WRONG — only B3 gets the formula, B4:B14 stay empty
ws["B3"] = "=VLOOKUP(A3, Data!A:C, 3, FALSE)"

# RIGHT — write to every cell
for r in range(3, 15):
    ws[f"B{r}"] = f"=VLOOKUP(A{r}, Data!$A:$C, 3, FALSE)"
```

**Also avoid these Excel 365-only functions** (evaluate to `#NAME?` in LibreOffice):

| Category | Forbidden Functions |
|---|---|
| Lambda functions | `LET`, `LAMBDA`, `MAP`, `REDUCE`, `SCAN`, `MAKEARRAY`, `BYROW`, `BYCOL` |
| Text functions | `TEXTSPLIT`, `TEXTBEFORE`, `TEXTAFTER`, `VALUETOTEXT`, `ARRAYTOTEXT` |
| Other Excel-only | `STOCKHISTORY`, `IMAGE`, `CHOOSECOLS`, `CHOOSEROWS`, `TAKE`, `DROP`, `EXPAND`, `WRAPCOLS`, `WRAPROWS`, `TOCOL`, `TOROW`, `VSTACK`, `HSTACK` |

## Safe Functions (single-cell, LibreOffice compatible)

These return one value per cell and work correctly:
- **Math**: `SUM`, `SUMIF`, `SUMIFS`, `SUMPRODUCT`, `AVERAGE`, `AVERAGEIF`, `MIN`, `MAX`, `ROUND`, `ABS`, `MOD`, `INT`
- **Lookup**: `VLOOKUP`, `HLOOKUP`, `XLOOKUP`, `XMATCH`, `INDEX`, `MATCH`, `OFFSET`, `INDIRECT`
- **Logic**: `IF`, `IFS`, `AND`, `OR`, `NOT`, `IFERROR`, `IFNA`
- **Count**: `COUNT`, `COUNTA`, `COUNTIF`, `COUNTIFS`, `COUNTBLANK`
- **Text**: `LEFT`, `RIGHT`, `MID`, `LEN`, `TRIM`, `UPPER`, `LOWER`, `CONCATENATE`, `TEXT`, `VALUE`, `FIND`, `SEARCH`, `SUBSTITUTE`, `REPT`
- **Date**: `DATE`, `YEAR`, `MONTH`, `DAY`, `TODAY`, `NOW`, `DATEDIF`, `EDATE`, `EOMONTH`, `WEEKDAY`, `NETWORKDAYS`
- **Other**: `ROW`, `COLUMN`, `ROWS`, `COLUMNS`, `LARGE`, `SMALL`, `RANK`, `PERCENTILE`

## CSE Array Formulas — BANNED

Formulas that pass `IF()` over a range into an aggregate function require Ctrl+Shift+Enter (CSE) array entry in Excel. openpyxl **cannot** write the CSE flag, so LibreOffice evaluates them as scalar — only the first cell is checked, and the rest return None.

**Never write these patterns as formulas:**
- `SMALL(IF(...), n)` — nth value matching a condition
- `MAX(IF(...))` / `MIN(IF(...))` — conditional max/min
- `INDEX(SMALL(IF(...)))` — lookup by nth match
- `LOOKUP(2, 1/(condition), range)` — last-match trick
- `AGGREGATE(...)` with array arguments
- Any formula where `IF()` evaluates over a range inside another function

**Always compute these in Python and write the value directly.** There is no workaround — these will silently return None or wrong results if written as formulas.

## When to Use Python vs Formulas

| Task | Approach |
|---|---|
| **Filtering rows** to populate a sheet | **Python**: read source rows, write matching rows with openpyxl |
| **Sorting data** into a sheet | **Python**: sort in Python, write sorted rows |
| **Listing unique values** into cells | **Python**: deduplicate in Python, write values |
| **Populating a sheet from another sheet** | **Python**: read source, filter/transform, write row by row |
| **Per-cell calculations** (sums, averages, ratios) | **Formula**: `=SUM(...)`, `=A1/B1`, etc. |
| **Per-cell conditional values** | **Formula**: `=IF(...)`, `=IFERROR(...)` |
| **Per-cell lookups** | **Formula**: `=VLOOKUP(...)` or `=INDEX(MATCH(...))` |
| **Conditional formatting** | **openpyxl**: use `openpyxl.formatting.rule` |

**The rule**: if the result fills multiple cells (filtering, sorting, listing), do it in Python and write the values. If the result is one value per cell (sum, lookup, conditional), use a formula.

## Important: Preserve Existing Data Order

When a workbook already has sheets with data that you need to update (e.g., Paid/NotPaid sheets with filtered records), **do not clear and rebuild from scratch**. Instead:
1. Read the existing sheet to understand its current content and ordering
2. Apply the minimum changes needed to satisfy the instruction
3. If you must rebuild, preserve the original row ordering

---

# Spreadsheet Operations Guide

## Output Standards

### Typography

Every generated workbook should default to a clean, widely available typeface — Arial or Calibri work well. Override only when the user specifies a preference or when an existing file already uses a different font family.

### Formula Integrity

Fix **unintentional** calculation errors (`#REF!`, `#DIV/0!`, `#NAME?`). A `#NAME?` error almost always means you used an Excel-only function — replace it with Python logic. When the instruction asks for a **formula**, write the formula — do not compute the result in Python and hardcode it. The formula may naturally produce `#VALUE!` or `#N/A` for some inputs, and that is correct behavior. Bypassing the formula with Python-computed values hides errors the formula would legitimately produce.

### Respecting Existing Files

When a user provides a template or partially completed workbook, treat its layout, fonts, colors, and naming conventions as authoritative. Mimic the existing patterns rather than overwriting them with defaults from this guide.

## Preserve Exact Formatting from Source Data

Before writing any label, category, or text value, **read 3-5 existing values in the target column** to learn the convention. Then match it exactly:
- ALL CAPS columns stay ALL CAPS — never title-case or lowercase them (e.g., `ACME` stays `ACME`, not `Acme`).
- Abbreviated values stay abbreviated — never expand them (e.g., `Wed` stays `Wed`, not `Wednesday`).
- Plural headers stay plural — don't drop the trailing "s" (e.g., `Total Expenses`, not `Total Expense`).
- Typos, extra spaces, and odd formatting in existing data are intentional — reproduce them verbatim, don't correct them.
- Preserve trailing spaces, padding characters, and delimited position strings exactly as they appear in the source.
- When writing fallback/default values in IFERROR formulas, match the casing convention of surrounding data.

## Never Write VBA or Explanation Sheets

Your task is always to produce a modified spreadsheet file with the actual data transformation applied. Even if the instruction says "write a formula", "use VBA", or "how would you do this", you must:

1. Implement the transformation directly using Python/openpyxl in the sandbox
2. Write the computed results or formulas into the appropriate cells
3. Never create sheets named "VBA_Code", "Answer", "Explanation", or similar
4. Never insert VBA macro text or step-by-step explanations into cells

If the instruction describes a task (delete rows, filter data, sort, etc.), do it — don't document how to do it.

---

## Choosing the Right Python Library

Two libraries cover the vast majority of spreadsheet tasks:

**pandas** is ideal when the job is fundamentally about data: reading a file, filtering rows, computing aggregates, pivoting, merging datasets, or exporting a cleaned table.

**openpyxl** is the right pick when the job involves presentation-layer concerns: cell-level formatting, conditional coloring, named ranges, embedded formulas, merged cells, or when you need to surgically edit a workbook without disturbing its existing structure.

For many tasks you'll use both — pandas to wrangle the data, openpyxl to place it into a formatted workbook.

---

## Reading & Analyzing Spreadsheet Data

```python
import pandas as pd

# Pull in the first sheet
df = pd.read_excel("quarterly_data.xlsx")

# Pull in every sheet as a dictionary of DataFrames
sheets = pd.read_excel("quarterly_data.xlsx", sheet_name=None)

# Quick inspection
df.head()
df.info()
df.describe()

# Export after transformations
df.to_excel("cleaned_output.xlsx", index=False)
```

### Tips for robust reads

- Pin column types to avoid silent misinterpretation: `pd.read_excel("f.xlsx", dtype={"account_id": str})`
- Limit memory usage on wide files by selecting only needed columns: `usecols=["Date", "Amount", "Category"]`
- Parse date columns at read time: `parse_dates=["transaction_date"]`

---

## Formulas vs Python — The Rule

Use LibreOffice-compatible Excel formulas for **per-cell calculations** that reference other cells: sums, averages, lookups, conditional logic. These stay live when the spreadsheet is opened.

Use Python to **write data directly** when the task involves filtering, sorting, deduplicating, or populating one sheet from another. Write the values row by row with openpyxl — don't try to use dynamic array formulas.

### Per-cell formulas (good)

```python
# Totals, ratios, conditionals — these work in LibreOffice
ws["F25"] = "=SUM(F3:F24)"
ws["G10"] = "=G8/G6"
ws["H5"] = '=IF(D5>100, "High", "Low")'
ws["B10"] = "=VLOOKUP(A10, Sheet2!A:C, 3, FALSE)"
```

### Robust lookups

Lookup formulas frequently fail due to trailing spaces or type mismatches in keys. Wrap lookup keys with `TRIM()` to handle whitespace:
```python
ws["B10"] = "=VLOOKUP(TRIM(A10), Sheet2!A:C, 3, FALSE)"
```
Do **not** wrap lookups in `IFERROR` by default — if the existing column already contains `#N/A` errors, those are intentional data. Only add `IFERROR` when the instruction explicitly asks to suppress errors.

### Filtering/populating sheets (do in Python)

```python
# Instead of =FILTER(...), iterate and write rows directly
from openpyxl import load_workbook

wb = load_workbook("input.xlsx")
src = wb["Source"]
dest = wb["Filtered"]

# Copy headers
for c in range(1, src.max_column + 1):
    dest.cell(1, c).value = src.cell(1, c).value

# Filter and write matching rows
out_row = 2
for r in range(2, src.max_row + 1):
    date_val = src.cell(r, 4).value
    if isinstance(date_val, datetime) and 2019 <= date_val.year <= 2023:
        for c in range(1, src.max_column + 1):
            dest.cell(out_row, c).value = src.cell(r, c).value
        out_row += 1

wb.save("output.xlsx")
```

---

## End-to-End Workflow

1. **Pick your tool.** Data wrangling → pandas. Formatting and formulas → openpyxl. Often both.
2. **Open or create** the workbook.
3. **Populate and style** — write data, insert formulas, apply formatting.
4. **Save** the `.xlsx` file.
5. **Verify formulas** — import `formula_eval` and call `evaluate()` to compute every formula in memory and check for errors. Fix any issues and re-verify until the report comes back clean.
6. **Validate with recalculate()** — If you wrote formulas, call `recalculate(file_path)` to evaluate them, then re-open with `load_workbook(file_path, data_only=True)` and spot-check key cells. If you find `None` where a value is expected, your formula didn't evaluate — replace it with a Python-computed value. If you find `#NAME?`, you used an unsupported function — replace with Python logic.

**IMPORTANT: Never call `recalculate()` and `SUBMIT()` in the same turn.**
The workflow must be:
- **Turn N**: Save the file, call `recalculate()`, re-open with
  `data_only=True`, and verify all modified cells. Fix any issues.
- **Turn N+1**: Only call `SUBMIT()` — nothing else.
Do not submit until all cells have been verified. Cells showing None
or error strings (#NAME?, #VALUE!, #REF!) must be fixed first —
typically by replacing the formula with a Python-computed value.

---

## Building a New Workbook

```python
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Header row
headers = ["Quarter", "Revenue", "COGS", "Gross Profit"]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, size=11, name="Arial")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center")

# Sample data
ws.append(["Q1", 150000, 90000])
ws.append(["Q2", 175000, 100000])

# Gross profit formula for each row
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=4).value = f"=B{r}-C{r}"

# Totals
total_row = ws.max_row + 1
ws.cell(row=total_row, column=1, value="Total")
ws.cell(row=total_row, column=2).value = f"=SUM(B2:B{total_row - 1})"
ws.cell(row=total_row, column=3).value = f"=SUM(C2:C{total_row - 1})"
ws.cell(row=total_row, column=4).value = f"=SUM(D2:D{total_row - 1})"

# Adjust column widths
for col_letter in ["A", "B", "C", "D"]:
    ws.column_dimensions[col_letter].width = 16

wb.save("quarterly_summary.xlsx")
```

---

## Modifying an Existing Workbook

```python
from openpyxl import load_workbook

wb = load_workbook("quarterly_summary.xlsx")
ws = wb["Summary"]

# Walk through every sheet
for name in wb.sheetnames:
    sheet = wb[name]
    print(f"Processing: {name}, rows={sheet.max_row}")

# Update a value
ws["B2"] = 162000

# Structural changes
ws.insert_rows(3)          # push row 3 down, insert blank
ws.delete_cols(5)          # remove column E

# Add a new tab
notes = wb.create_sheet("Notes")
notes["A1"] = "Model last updated 2026-04-01"

wb.save("quarterly_summary_v2.xlsx")
```

---

## Verifying Formulas with `formula_eval`

openpyxl writes formula strings into cells but does **not** evaluate them. To verify that every formula resolves correctly, use the `formula_eval` module which computes all values in pure Python — no external applications needed.

```python
from formula_eval import evaluate

report = evaluate("quarterly_summary.xlsx")

if report["ok"]:
    print("All formulas clean")
else:
    for token, info in report["breakdown"].items():
        print(f"{token}: {info['count']} error(s)")
        for cell in info["cells"]:
            print(f"  {cell}")
```

You can also inspect individual computed values programmatically:

```python
report = evaluate("quarterly_summary.xlsx")

# Check a specific cell's computed result
assert report["computed"]["Summary!D4"] == 325000
```

**On large workbooks, always pass `cells` to avoid sandbox hangs.** Evaluating every formula in a workbook with hundreds of rows can freeze the sandbox. Instead, pass only the cells you wrote:

```python
# SLOW / MAY HANG — evaluates every formula in the workbook
report = evaluate("big_file.xlsx")

# FAST — only evaluates these cells and their dependencies
report = evaluate("big_file.xlsx", cells=["Sheet1!B5", "Sheet1!B6", "Sheet1!B7"])
```

### Understanding the report

The `evaluate()` function returns a dictionary with these keys:

| Key | Type | Description |
|---|---|---|
| `ok` | bool | `True` when every formula resolved without an error token |
| `formulas` | int | Total formula cells found in the workbook |
| `errors` | int | How many cells evaluated to an Excel error |
| `breakdown` | dict | Each error token mapped to `{"count": N, "cells": [...]}` |
| `computed` | dict | Every evaluated cell as `"Sheet!Cell": value` — useful for assertions |

When errors appear: fix the offending formulas in openpyxl, save again, and re-run `evaluate()` until `errors` reaches zero. A `#NAME?` error means you used a function LibreOffice doesn't support — replace it with Python logic.

### Ensuring spreadsheet apps recalculate on open

```python
from formula_eval import ensure_recalc_on_open

ensure_recalc_on_open("quarterly_summary.xlsx")
```

This sets the `fullCalcOnLoad` flag in the workbook metadata, guaranteeing that LibreOffice will do a fresh computation pass when the file is opened.

---

## Formula Debugging Checklist

### Before you build

- Spot-check two or three cell references to make sure they resolve to the values you expect.
- Confirm your column-number-to-letter mapping — column 64 is `BL`, not `BK`.
- Account for the index offset between pandas (0-based) and Excel (1-based). A DataFrame's row index 5 lands on Excel row 6.

### Common traps

- **`#NAME?` errors**: you used an Excel-only function. Replace with Python logic or a LibreOffice-compatible formula.
- **`#VALUE!` from `SUMIF`/`SUMIFS` with array math**: `SUMIF` does NOT support computed arrays (e.g., multiplying two ranges) as the sum_range. Use `SUMPRODUCT` for any criteria-based calculation involving array expressions.
- **Circular references**: never write a formula into a cell that references itself. If you need to replace a cell's value with a computed result, read the original value into Python first, then write the result back as a literal value — not as a formula referencing the same cell.
- **Null values**: screen with `pd.notna()` before writing to cells.
- **Wide datasets**: fiscal-year data frequently sits beyond column 50 — don't stop scanning early.
- **Duplicate matches**: when searching for a label, verify you've found the right occurrence if it appears more than once.
- **Division by zero**: guard every divisor — wrap in `IFERROR` or check the denominator cell.
- **Broken references**: after inserting or deleting rows/columns, confirm existing formulas still point where they should.
- **Multi-sheet links**: the correct syntax is `SheetName!A1`; forgetting the sheet name prefix silently targets the active sheet.
- **Array formulas with literal braces**: Never write `{=FORMULA}` with curly braces — openpyxl stores the braces as literal text, not as an array formula marker. The cell will display the formula text instead of computing a result. Either write `=FORMULA` without braces (works for most cases in LibreOffice), or compute the result in Python and write the value directly.
- **Datetime vs time-only**: To store time-only values (no date component), write `=MOD(cell,1)` to extract the time fraction, or compute the time in Python and write a float (e.g., 0.75 for 18:00). Apply a time number format like `h:mm AM/PM`. Do not copy a full datetime cell when only the time is expected.
- **Merged cells when restructuring**: openpyxl only stores the value in the top-left cell of a merged range; all others read as `None`. Before copying or restructuring rows, check `ws.merged_cells.ranges` and unmerge first. After restructuring, re-merge if the output requires it. Do not duplicate the top-left value across all cells of a former merge.
- **Sheet name accuracy**: When the instruction names a target sheet, create it with the exact name given (including capitalization and spaces). After saving, verify the output workbook contains the expected sheet. If the instruction's sheet name doesn't exist in the input file, create a new sheet — do not write to an existing sheet with a different name.
- **LibreOffice requires `_xlfn.` prefixes for modern functions**: Functions such as `IFNA`, `IFS`, `SWITCH`, `TEXTJOIN`, `AGGREGATE`, `MAXIFS`, `MINIFS`, and `CONCAT` must be written as `=_xlfn.IFNA(...)`, etc. Without the prefix LibreOffice recalculates them to `#NAME?` even if Excel accepted them.
- **`#VALUE!` from mixed types in ranges**: Functions like `ISNUMBER(SEARCH(...))`, `SUMPRODUCT`, and `MATCH` produce `#VALUE!` when a range contains a mix of numbers, strings, and blanks. Clean the data in Python first or guard with `IFERROR`. If computing over mixed types, prefer doing the logic in Python and writing the result as a value.
- **Avoid full-column refs when the data footprint is small**: Prefer bounded ranges that match the actual data extent (e.g., use `ws.max_row` to build `$A$2:$A${max_row}`). Full-column references like `$A:$A` force the sandbox formula verifier to iterate ~1M rows, which hangs. Only fall back to `$A:$A` when the instruction clearly expects the column to stay open-ended.
- **Match the column's stored semantics before writing**: Inspect nearby rows to learn how booleans, empty slots, and categorical labels are encoded. Write using those exact tokens — e.g., cast booleans with `int(...)` when the column stores 1/0 flags, only emit `"-"` if the sheet uses that placeholder, and preserve abbreviations and casing exactly as they appear.
- **Keep intentional error/blank guardrails intact**: When the spec expects a control cell to stay blank or show `#VALUE!`/`#REF!`, reproduce that state. Don't "fix" guardrail formulas or populate header rows that are meant to remain empty.
- **Audit the target range before overwriting**: Before writing across a range, scan existing cells for total/summary formulas (e.g., `=SUM(...)`, `=AVERAGE(...)`), header rows, and blank-vs-zero semantics. Preserve pre-existing formulas you didn't create. Guard non-data rows so header or separator rows stay empty.
- **SUM on text-formatted numbers returns 0**: If source cells contain numbers stored as text, `SUM()` silently returns 0. Either coerce with `VALUE()` inside `SUMPRODUCT`, or read the values in Python and write numeric literals.
- **INDEX returning a blank cell yields 0 in numeric context**: Wrap with `IF(INDEX(...)="","",INDEX(...))` to preserve true emptiness when the output should be blank.
- **Validate lookup offsets against known values**: Off-by-one in MATCH/INDEX column offsets silently returns the adjacent wrong column. After writing a lookup formula, compute the expected result for 2-3 rows in Python and compare.
- **Aggressive string normalization merges distinct keys**: Stripping whitespace or lowercasing lookup keys can merge entries that differ only by trailing spaces or casing. Inspect actual key values before deciding how to normalize.

### Verify formulas before saving

After writing formulas to cells, you cannot rely on openpyxl to evaluate them. Instead:

1. **Compute expected values in Python** for at least 3 representative rows: the first data row, a middle row, and the last row. Include edge cases (zero values, empty cells, boundary conditions).
2. **Compare against the instruction's examples** if any are provided.
3. **Check formula references**: verify that row/column anchoring (`$A$1` vs `A1`) is correct. A formula that works in row 2 may break in row 50 if relative references shift incorrectly.
4. **Watch for cumulative formulas**: running totals, rolling averages, and phased calculations are especially error-prone. Manually trace 3 rows to confirm the accumulation logic.

---

## openpyxl Essentials

- Rows and columns are **1-indexed**: `cell(row=1, column=1)` is `A1`.
- To read previously computed values without formulas, open with `data_only=True`. **Caution**: saving a workbook opened this way permanently replaces formulas with their last-calculated values.
- For very large files, `read_only=True` (reading) and `write_only=True` (writing) dramatically reduce memory consumption.
- Formulas written by openpyxl are stored as text — use `formula_eval.evaluate()` to verify their computed values in Python.
- **NEVER use `ws.delete_rows()` or `ws.insert_rows()`** — these are O(n²) in WASM and will hang on sheets with more than a few hundred rows. Instead, read all rows into a list, filter/transform in Python, then write back by overwriting cells in place. Clear leftover rows at the bottom by setting cell values to `None`.

## pandas Essentials

- Force column types at load time to prevent silent coercion: `dtype={"zip_code": str}`.
- Narrow wide files by selecting only relevant columns: `usecols=["A", "D", "F"]`.
- Let pandas handle date parsing: `parse_dates=["order_date"]`.

---

## Code Style Notes

When writing Python for spreadsheet tasks, favor brevity: short variable names, minimal inline commentary, no diagnostic print statements unless you're actively debugging. The code is a means to an end — the workbook is the deliverable.

Inside the workbook itself, be generous with documentation: annotate complex formulas with cell comments, tag every hardcoded input with its source, and add section labels so future readers can navigate the model without reverse-engineering it.
""",
    packages=["openpyxl", "pandas", "formulas"],
    modules={"formula_eval": str(_BUILTIN_MODULES_DIR / "formula_eval.py")},
    tools={"recalculate": recalculate},
)
