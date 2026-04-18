"""xlsx → PNG rendering pipeline for visual verification.

Converts a workbook to an image via LibreOffice (``soffice --convert-to pdf``)
then rasterizes the first page to PNG via Poppler's ``pdftoppm``. Used by the
spreadsheet skill's ``render`` tool to give sub-LMs a rasterized view of a
sheet they can inspect with ``dspy.Image``.

Typical usage::

    from spreadsheet_rlm.render import render

    # Whole active sheet
    result = render("model_output.xlsx")

    # Just a cell range (sets a print area on a temp copy before PDF export)
    result = render("model_output.xlsx", cell_range="J3:N5")

    # Sheet-qualified range
    result = render("model_output.xlsx", cell_range="Sheet1!J3:N5")
"""

from __future__ import annotations

import base64
import logging
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .recalculate import _find_libreoffice

log = logging.getLogger("spreadsheet_rlm.render")

_RENDER_TIMEOUT_SECONDS = 120.0


@dataclass
class RenderResult:
    """Outcome of a :func:`render` call."""

    png_bytes: bytes
    source_xlsx: Path


def render(
    path: str | Path,
    cell_range: str | None = None,
    sheet_name: str | None = None,
) -> RenderResult:
    """Render *path* to a PNG of the first page.

    Args:
        path: Path to the xlsx workbook.
        cell_range: Optional cell range to restrict the render to, e.g.
            ``"J3:N5"`` or ``"Sheet1!J3:N5"``. When given, the workbook is
            copied to a temp file, a print area is set on the target sheet,
            and LibreOffice exports only that region. This is the right
            tool for "show me what this specific area looks like" — the
            default whole-sheet render only captures the first printed
            page, which misses columns/rows outside the first page on
            wide or tall sheets.
        sheet_name: Sheet to set the print area on when *cell_range* is
            given. Ignored if *cell_range* already contains a
            ``Sheet!range`` prefix. Defaults to the workbook's active
            sheet.

    Raises:
        FileNotFoundError: if *path* does not exist.
        ValueError: if *sheet_name* is given but not found in the workbook,
            or if *cell_range* is malformed.
        RuntimeError: if LibreOffice or ``pdftoppm`` is unavailable, or if
            either subprocess fails or produces no output.
    """
    src = Path(path).resolve()
    if not src.is_file():
        raise FileNotFoundError(f"workbook not found: {src}")

    soffice = _find_libreoffice()
    if not soffice:
        raise RuntimeError("LibreOffice (soffice) not found on PATH")

    pdftoppm = _find_pdftoppm()
    if not pdftoppm:
        raise RuntimeError("pdftoppm (from Poppler) not found on PATH")

    with tempfile.TemporaryDirectory() as tmpname:
        tmp = Path(tmpname)
        profile = tmp / "profile"
        profile.mkdir()
        pdf_dir = tmp / "pdf"
        pdf_dir.mkdir()
        png_dir = tmp / "png"
        png_dir.mkdir()

        xlsx_to_export = src
        if cell_range is not None:
            xlsx_to_export = tmp / src.name
            _apply_print_area(src, xlsx_to_export, cell_range, sheet_name)

        lo = subprocess.run(
            [
                soffice,
                "--headless",
                "--calc",
                "--convert-to",
                "pdf",
                "--outdir",
                str(pdf_dir),
                f"-env:UserInstallation=file://{profile}",
                str(xlsx_to_export),
            ],
            capture_output=True,
            text=True,
            timeout=_RENDER_TIMEOUT_SECONDS,
        )
        if lo.returncode != 0:
            msg = lo.stderr.strip() or lo.stdout.strip() or "unknown error"
            raise RuntimeError(f"soffice --convert-to pdf failed: {msg}")

        pdf_candidates = list(pdf_dir.glob("*.pdf"))
        if not pdf_candidates:
            raise RuntimeError("soffice produced no pdf output")
        pdf_path = pdf_candidates[0]

        png_stem = png_dir / "page"
        rc = subprocess.run(
            [
                pdftoppm,
                "-png",
                "-f", "1",
                "-l", "1",
                str(pdf_path),
                str(png_stem),
            ],
            capture_output=True,
            text=True,
            timeout=_RENDER_TIMEOUT_SECONDS,
        )
        if rc.returncode != 0:
            msg = rc.stderr.strip() or rc.stdout.strip() or "unknown error"
            raise RuntimeError(f"pdftoppm failed: {msg}")

        png_candidates = sorted(png_dir.glob("page*.png"))
        if not png_candidates:
            raise RuntimeError("pdftoppm produced no png output")
        png_bytes = png_candidates[0].read_bytes()

    return RenderResult(png_bytes=png_bytes, source_xlsx=src)


def render_to_data_uri(
    path: str | Path,
    cell_range: str | None = None,
    sheet_name: str | None = None,
) -> str:
    """Render *path* and return a ``data:image/png;base64,...`` URI."""
    result = render(path, cell_range=cell_range, sheet_name=sheet_name)
    b64 = base64.b64encode(result.png_bytes).decode()
    return f"data:image/png;base64,{b64}"


def _apply_print_area(
    src: Path,
    dst: Path,
    cell_range: str,
    sheet_name: str | None,
) -> None:
    """Copy *src* to *dst* and set its print area so LibreOffice exports only that region.

    Parses sheet-qualified ranges like ``"Sheet1!J3:N5"`` or plain
    ``"J3:N5"``. If no sheet is specified (neither via the prefix nor via
    *sheet_name*), defaults to the workbook's active sheet.
    """
    if "!" in cell_range:
        sheet_part, range_part = cell_range.split("!", 1)
        parsed_sheet = sheet_part.strip().strip("'")
        parsed_range = range_part.strip()
    else:
        parsed_sheet = sheet_name
        parsed_range = cell_range.strip()

    if not parsed_range:
        raise ValueError(f"empty cell range in: {cell_range!r}")

    shutil.copy2(src, dst)
    wb = load_workbook(dst)
    try:
        if parsed_sheet is None:
            ws = wb.active
        elif parsed_sheet in wb.sheetnames:
            ws = wb[parsed_sheet]
        else:
            raise ValueError(
                f"sheet {parsed_sheet!r} not in workbook (sheets: {wb.sheetnames})"
            )
        ws.print_area = parsed_range
        wb.save(dst)
    finally:
        wb.close()


def _find_pdftoppm() -> str | None:
    """Locate the ``pdftoppm`` binary, or return ``None`` if absent."""
    candidates = [
        "pdftoppm",
        "/opt/homebrew/bin/pdftoppm",
        "/usr/bin/pdftoppm",
        "/usr/local/bin/pdftoppm",
    ]
    for c in candidates:
        path = shutil.which(c)
        if path:
            return path
        if Path(c).is_file():
            return c
    return None
