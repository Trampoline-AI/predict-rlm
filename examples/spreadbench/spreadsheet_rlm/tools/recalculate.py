"""Formula recalculation pipeline for xlsx files.

Resolves stale formula values inside a workbook by running it through a
two-stage pipeline:

1. **Python `formulas` library** — computes every formula in memory
   using a pure-Python dependency graph. Executed in an isolated child
   process with a hard timeout so pathological graphs cannot wedge the
   host run, and handles a surprisingly wide range of Excel functions including
   `TEXTJOIN` and other modern built-ins.
2. **LibreOffice headless** — ``soffice --headless --convert-to xlsx``
   as a fallback for formulas the Python library can't evaluate.

The pipeline picks whichever candidate resolved the most formula cells,
with the untouched baseline always in the running so the call is
strictly additive: the winning file has at least as many resolved
formula cells as the input. This mirrors the "zero regressions" property
the spreadbench eval relies on — a recalc never turns a passing test
into a failing one by destroying already-cached values.

Typical usage::

    from spreadsheet_rlm.tools.recalculate import recalculate

    result = recalculate("model_output.xlsx")
    # result.source      -> "baseline" | "formulas" | "libreoffice"
    # result.resolved    -> number of resolved formula cells in the winner
    # result.errors      -> list of per-stage failure messages (if any)
"""

from __future__ import annotations

import logging
import multiprocessing as mp
import os
import re
import shutil
import subprocess
import tempfile
import time
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import openpyxl

from predict_rlm.telemetry import TelemetryContext, current_telemetry_context

try:
    import formulas as _formulas_lib
except ImportError:
    _formulas_lib = None

# Suppress openpyxl's benign UserWarnings about dropped xlsx features
# (e.g. "Data Validation extension is not supported and will be removed",
# "Conditional Formatting extension is not supported and will be removed",
# "Title is more than 31 characters..."). These fire frequently on the
# spreadbench dataset — they're informational about what openpyxl won't
# round-trip on save, and have no effect on our scoring. Scoped to the
# openpyxl module only so we don't hide warnings from anything else.
warnings.filterwarnings("ignore", category=UserWarning, module=r"openpyxl\..*")

log = logging.getLogger("spreadsheet_rlm.tools.recalculate")

_FORMULAS_TQDM_PATCHED = False


class _QuietTqdm:
    """No-op stand-in for ``tqdm.tqdm`` used by the formulas library.

    Mirrors the QuietTqdm pattern from ``formulas/cli.py`` — when monkey
    patched into the tqdm reference inside ``formulas.excel``, every
    progress bar call becomes a no-op. Used to keep the formulas
    library's internal progress bars from interleaving with the eval
    loop's own tqdm output during recalculation.
    """

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        pass

    def __enter__(self) -> "_QuietTqdm":
        return self

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> bool:
        return False

    def update(self, *args: Any, **kwargs: Any) -> None:
        return None

    def refresh(self) -> None:
        return None

    total = 0


def _silence_formulas_tqdm() -> None:
    global _FORMULAS_TQDM_PATCHED
    if _FORMULAS_TQDM_PATCHED or _formulas_lib is None:
        return
    _formulas_lib.ExcelModel.complete.__globals__["tqdm"].tqdm = _QuietTqdm
    _FORMULAS_TQDM_PATCHED = True

_ERROR_TOKENS = frozenset(
    ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
)

# Keys returned by the formulas library look like:
#   "'[workbook.xlsx]SHEETNAME'!A1"
# The sheet name is always upper-cased by the library.
_CELL_KEY_RE = re.compile(r"'?\[.*?\](.+?)'?!([A-Z]+\d+)$")

_FORMULAS_TIMEOUT_SECONDS = 30.0
_FORMULAS_TERMINATE_GRACE_SECONDS = 2.0

Source = Literal["baseline", "formulas", "libreoffice"]


@dataclass
class RecalcResult:
    """Outcome of a :func:`recalculate` call."""

    source: Source
    resolved: int
    total_formulas: int
    baseline_resolved: int
    formulas_resolved: int | None = None
    libreoffice_resolved: int | None = None
    errors: list[str] = field(default_factory=list)


def recalculate(
    path: str | Path,
    *,
    telemetry_context: TelemetryContext | None = None,
) -> RecalcResult:
    """Recalculate formulas in *path*, writing the best candidate back in place.

    The pipeline snapshots the baseline workbook, runs the Python
    `formulas` library, and falls back to LibreOffice headless if any
    formulas were left unresolved. The candidate with the most resolved
    formula cells wins; baseline is always a candidate so the file is
    never downgraded.

    Returns a :class:`RecalcResult`. Ties prefer the least destructive
    source in the order ``baseline > libreoffice > formulas`` —
    baseline changes nothing, LibreOffice keeps formula strings intact,
    and the formulas library flattens computed formulas to literal
    values.
    """
    telemetry_context = telemetry_context or current_telemetry_context()
    start_ns = time.time_ns()
    src = Path(path).resolve()
    if not src.is_file():
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate",
            start_time_unix_nano=start_ns,
            status={"code": "ERROR", "message": "workbook not found"},
            attributes={"failure.class": "evaluator_limitation"},
        )
        raise FileNotFoundError(f"workbook not found: {src}")

    targets = _formula_targets(src)
    total = len(targets)

    if total == 0:
        result = RecalcResult(
            source="baseline",
            resolved=0,
            total_formulas=0,
            baseline_resolved=0,
        )
        _write_recalc_result_span(telemetry_context, start_ns, result, branch="no_formulas")
        return result

    baseline_resolved = _count_resolved(src, targets)
    log.debug("baseline: %d/%d formula cells resolved", baseline_resolved, total)

    if baseline_resolved == total:
        result = RecalcResult(
            source="baseline",
            resolved=baseline_resolved,
            total_formulas=total,
            baseline_resolved=baseline_resolved,
        )
        _write_recalc_result_span(
            telemetry_context,
            start_ns,
            result,
            branch="baseline_already_resolved",
        )
        return result

    errors: list[str] = []
    candidates: dict[Source, tuple[Path, int]] = {}
    formulas_resolved: int | None = None
    libreoffice_resolved: int | None = None

    with tempfile.TemporaryDirectory() as tmpname:
        tmp = Path(tmpname)

        baseline_copy = tmp / "baseline.xlsx"
        shutil.copy2(src, baseline_copy)
        candidates["baseline"] = (baseline_copy, baseline_resolved)

        formulas_out = tmp / "formulas.xlsx"
        try:
            _recalc_with_formulas(src, formulas_out, telemetry_context=telemetry_context)
            formulas_resolved = _count_resolved(formulas_out, targets)
            candidates["formulas"] = (formulas_out, formulas_resolved)
            log.debug("formulas: %d/%d resolved", formulas_resolved, total)
        except Exception as e:
            errors.append(f"formulas: {e}")
            log.debug("formulas pipeline failed: %s", e)

        # Fast path: if the Python library resolved everything, skip the
        # LibreOffice subprocess entirely — that's the whole point of
        # running formulas first.
        if formulas_resolved == total:
            _replace(candidates["formulas"][0], src)
            result = RecalcResult(
                source="formulas",
                resolved=formulas_resolved,
                total_formulas=total,
                baseline_resolved=baseline_resolved,
                formulas_resolved=formulas_resolved,
                libreoffice_resolved=None,
                errors=errors,
            )
            _write_recalc_result_span(telemetry_context, start_ns, result, branch="formulas")
            return result

        lo_out = tmp / "libreoffice.xlsx"
        try:
            _recalc_with_libreoffice(src, lo_out, telemetry_context=telemetry_context)
            libreoffice_resolved = _count_resolved(lo_out, targets)
            candidates["libreoffice"] = (lo_out, libreoffice_resolved)
            log.debug("libreoffice: %d/%d resolved", libreoffice_resolved, total)
        except Exception as e:
            errors.append(f"libreoffice: {e}")
            log.debug("libreoffice pipeline failed: %s", e)

        winner = _pick_winner(candidates)
        winner_path, winner_count = candidates[winner]
        if winner != "baseline":
            _replace(winner_path, src)

    result = RecalcResult(
        source=winner,
        resolved=winner_count,
        total_formulas=total,
        baseline_resolved=baseline_resolved,
        formulas_resolved=formulas_resolved,
        libreoffice_resolved=libreoffice_resolved,
        errors=errors,
    )
    _write_recalc_result_span(telemetry_context, start_ns, result, branch=winner)
    return result


def _formula_targets(path: Path) -> list[tuple[str, str]]:
    """Return ``[(sheet_name, cell_coordinate)]`` for every formula cell."""
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    result: list[tuple[str, str]] = []
    try:
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for cell in row:
                    val = cell.value
                    if not (isinstance(val, str) and val.startswith("=")):
                        continue
                    # In read-only mode empty cells are EmptyCell instances
                    # that don't carry a coordinate attribute; skip defensively.
                    coord = getattr(cell, "coordinate", None)
                    if coord is not None:
                        result.append((name, coord))
    finally:
        wb.close()
    return result


def _count_resolved(path: Path, targets: list[tuple[str, str]]) -> int:
    """Count *targets* whose cached value in *path* is a non-error, non-null."""
    if not targets:
        return 0

    wb = openpyxl.load_workbook(path, data_only=True)
    resolved = 0
    try:
        for sheet_name, coord in targets:
            if sheet_name not in wb.sheetnames:
                continue
            val = wb[sheet_name][coord].value
            if val is None:
                continue
            if isinstance(val, str) and val in _ERROR_TOKENS:
                continue
            resolved += 1
    finally:
        wb.close()
    return resolved


def _run_formulas_worker(src: str, dst: str, send_conn: Any) -> None:
    try:
        _recalc_with_formulas_inprocess(Path(src), Path(dst))
        send_conn.send((True, ""))
    except Exception as e:
        send_conn.send((False, f"{type(e).__name__}: {e}"))
    finally:
        send_conn.close()


def _recalc_with_formulas(
    src: Path,
    dst: Path,
    *,
    telemetry_context: TelemetryContext | None = None,
) -> None:
    """Run formulas in a child process with a hard timeout."""
    if _formulas_lib is None:
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.formulas_worker",
            status={"code": "ERROR", "message": "formulas library not installed"},
            attributes={"failure.class": "evaluator_limitation"},
        )
        raise RuntimeError("formulas library not installed")

    start_ns = time.time_ns()
    context = mp.get_context("spawn")
    recv_conn, send_conn = context.Pipe(duplex=False)
    worker = context.Process(
        target=_run_formulas_worker,
        args=(str(src), str(dst), send_conn),
    )
    worker.start()
    _write_recalc_span(
        telemetry_context,
        "host_tool.recalculate.formulas_worker.start",
        start_time_unix_nano=start_ns,
        end_time_unix_nano=start_ns,
        attributes={
            "process.pid": worker.pid,
            "timeout.seconds": _FORMULAS_TIMEOUT_SECONDS,
        },
    )
    send_conn.close()

    worker.join(_FORMULAS_TIMEOUT_SECONDS)
    if worker.is_alive():
        killed = False
        worker.terminate()
        worker.join(_FORMULAS_TERMINATE_GRACE_SECONDS)
        if worker.is_alive():
            worker.kill()
            killed = True
            worker.join()
        recv_conn.close()
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.formulas_worker.timeout",
            start_time_unix_nano=start_ns,
            status={
                "code": "ERROR",
                "message": f"timed out after {_FORMULAS_TIMEOUT_SECONDS:.1f}s",
            },
            attributes={
                "process.pid": worker.pid,
                "process.exit_code": worker.exitcode,
                "process.killed": killed,
                "timeout.seconds": _FORMULAS_TIMEOUT_SECONDS,
                "failure.class": "host_tool_timeout_or_leak",
            },
        )
        raise RuntimeError(f"timed out after {_FORMULAS_TIMEOUT_SECONDS:.1f}s")

    ok = False
    message = f"worker exited without result (exitcode={worker.exitcode})"
    if recv_conn.poll():
        ok, message = recv_conn.recv()
    recv_conn.close()

    if not ok:
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.formulas_worker",
            start_time_unix_nano=start_ns,
            status={"code": "ERROR", "message": message},
            attributes={
                "process.pid": worker.pid,
                "process.exit_code": worker.exitcode,
            },
        )
        raise RuntimeError(message)
    if not dst.is_file():
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.formulas_worker",
            start_time_unix_nano=start_ns,
            status={"code": "ERROR", "message": "worker produced no output workbook"},
            attributes={
                "process.pid": worker.pid,
                "process.exit_code": worker.exitcode,
                "failure.class": "evaluator_exception",
            },
        )
        raise RuntimeError("worker produced no output workbook")
    _write_recalc_span(
        telemetry_context,
        "host_tool.recalculate.formulas_worker",
        start_time_unix_nano=start_ns,
        attributes={
            "process.pid": worker.pid,
            "process.exit_code": worker.exitcode,
        },
    )


def _recalc_with_formulas_inprocess(src: Path, dst: Path) -> None:
    """Run formulas in-process and write results to *dst*."""
    if _formulas_lib is None:
        raise RuntimeError("formulas library not installed")

    _silence_formulas_tqdm()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        model = _formulas_lib.ExcelModel().loads(str(src)).finish()
        solution = model.calculate()

    computed: dict[tuple[str, str], Any] = {}
    for key, value in solution.items():
        m = _CELL_KEY_RE.match(key)
        if not m:
            continue
        sheet_upper, cell_ref = m.group(1), m.group(2)
        scalar = _extract_scalar(value)
        if scalar is None:
            continue
        if isinstance(scalar, str) and scalar in _ERROR_TOKENS:
            continue
        computed[(sheet_upper, cell_ref)] = scalar

    wb = openpyxl.load_workbook(src)
    try:
        for sheet in wb.worksheets:
            key_sheet = sheet.title.upper()
            for row in sheet.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    new_value = computed.get((key_sheet, cell.coordinate))
                    if new_value is not None:
                        cell.value = new_value
        wb.save(dst)
    finally:
        wb.close()


def _recalc_with_libreoffice(
    src: Path,
    dst: Path,
    *,
    telemetry_context: TelemetryContext | None = None,
) -> None:
    """Recalculate *src* via ``soffice --headless`` and write it to *dst*."""
    soffice = _find_libreoffice()
    if not soffice:
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.libreoffice",
            status={"code": "ERROR", "message": "LibreOffice not found"},
            attributes={"failure.class": "evaluator_limitation"},
        )
        raise RuntimeError("LibreOffice not found")

    original_names: list[str] = []
    try:
        wb = openpyxl.load_workbook(src, read_only=True)
        original_names = list(wb.sheetnames)
        wb.close()
    except Exception:
        pass

    with tempfile.TemporaryDirectory() as tmpname:
        tmp = Path(tmpname)
        profile = tmp / "profile"
        profile.mkdir()
        outdir = tmp / "out"
        outdir.mkdir()
        cmd = [
            soffice,
            "--headless",
            "--calc",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(outdir),
            f"-env:UserInstallation=file://{profile}",
            str(src),
        ]
        start_ns = time.time_ns()
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.libreoffice.start",
            start_time_unix_nano=start_ns,
            end_time_unix_nano=start_ns,
            attributes={
                "process.executable": Path(soffice).name,
                "timeout.seconds": 120,
            },
        )
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=120,
            )
        except subprocess.TimeoutExpired as exc:
            _write_recalc_span(
                telemetry_context,
                "host_tool.recalculate.libreoffice.timeout",
                start_time_unix_nano=start_ns,
                status={"code": "ERROR", "message": "LibreOffice timed out"},
                attributes={
                    "process.executable": Path(soffice).name,
                    "timeout.seconds": 120,
                    "subprocess.stderr_tail_hash": _tail_hash(exc.stderr),
                    "subprocess.stderr_tail_len": len(_tail(exc.stderr)),
                    "failure.class": "host_tool_timeout_or_leak",
                },
            )
            raise
        if result.returncode != 0:
            msg = result.stderr.strip() or result.stdout.strip() or "unknown error"
            _write_recalc_span(
                telemetry_context,
                "host_tool.recalculate.libreoffice",
                start_time_unix_nano=start_ns,
                status={"code": "ERROR", "message": "LibreOffice failed"},
                attributes={
                    "process.executable": Path(soffice).name,
                    "process.exit_code": result.returncode,
                    "subprocess.stderr_tail_hash": _tail_hash(result.stderr),
                    "subprocess.stderr_tail_len": len(_tail(result.stderr)),
                    "failure.class": "evaluator_exception",
                },
            )
            raise RuntimeError(f"libreoffice failed: {msg}")
        produced = list(outdir.glob("*.xlsx"))
        if not produced:
            _write_recalc_span(
                telemetry_context,
                "host_tool.recalculate.libreoffice",
                start_time_unix_nano=start_ns,
                status={"code": "ERROR", "message": "LibreOffice produced no xlsx output"},
                attributes={
                    "process.executable": Path(soffice).name,
                    "process.exit_code": result.returncode,
                    "failure.class": "evaluator_exception",
                },
            )
            raise RuntimeError("libreoffice produced no xlsx output")
        shutil.copy2(produced[0], dst)
        _write_recalc_span(
            telemetry_context,
            "host_tool.recalculate.libreoffice",
            start_time_unix_nano=start_ns,
            attributes={
                "process.executable": Path(soffice).name,
                "process.exit_code": result.returncode,
                "subprocess.stderr_tail_hash": _tail_hash(result.stderr),
                "subprocess.stderr_tail_len": len(_tail(result.stderr)),
            },
        )

    if original_names:
        _restore_sheet_names(dst, original_names)


def _restore_sheet_names(path: Path, names: list[str]) -> None:
    """Rename sheets back to *names* if LibreOffice renamed them in place."""
    try:
        wb = openpyxl.load_workbook(path)
        if wb.sheetnames != names and len(wb.sheetnames) == len(names):
            for ws, orig in zip(wb.worksheets, names):
                if ws.title != orig:
                    ws.title = orig
            wb.save(path)
        wb.close()
    except Exception:
        pass


def _find_libreoffice() -> str | None:
    """Locate the LibreOffice binary, or return None if absent."""
    candidates = [
        "soffice",
        "libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
    ]
    for c in candidates:
        path = shutil.which(c)
        if path:
            return path
        if os.path.isfile(c):
            return c
    return None


def _pick_winner(candidates: dict[Source, tuple[Path, int]]) -> Source:
    """Pick the candidate with the most resolved cells.

    Iteration order ``baseline -> libreoffice -> formulas`` combined with a
    strict ``>`` comparison means ties are resolved in favour of the
    earlier (less destructive) candidate: baseline wins ties against
    everything, LibreOffice wins ties against formulas.
    """
    preference: list[Source] = ["baseline", "libreoffice", "formulas"]
    winner: Source = "baseline"
    winner_count = candidates["baseline"][1]
    for name in preference[1:]:
        if name not in candidates:
            continue
        count = candidates[name][1]
        if count > winner_count:
            winner = name
            winner_count = count
    return winner


def _extract_scalar(value: Any) -> Any:
    """Coerce a `formulas` library return value to a plain Python scalar."""
    try:
        import numpy as np
    except ImportError:
        np = None

    if hasattr(value, "value"):
        value = value.value
    if np is not None and isinstance(value, np.ndarray):
        if value.size == 0:
            return None
        value = value.flat[0]
    if np is not None:
        if isinstance(value, np.integer):
            return int(value)
        if isinstance(value, np.floating):
            return float(value)
    return value


def _replace(src: Path, dst: Path) -> None:
    """Overwrite *dst* with the contents of *src*, preserving metadata."""
    shutil.copy2(src, dst)


def _write_recalc_result_span(
    telemetry_context: TelemetryContext | None,
    start_ns: int,
    result: RecalcResult,
    *,
    branch: str,
) -> None:
    _write_recalc_span(
        telemetry_context,
        "host_tool.recalculate",
        start_time_unix_nano=start_ns,
        attributes={
            "spreadbench.recalculate.branch": branch,
            "spreadbench.recalculate.source": result.source,
            "spreadbench.recalculate.total_formulas": result.total_formulas,
            "spreadbench.recalculate.resolved": result.resolved,
            "spreadbench.recalculate.baseline_resolved": result.baseline_resolved,
            "spreadbench.recalculate.formulas_resolved": result.formulas_resolved,
            "spreadbench.recalculate.libreoffice_resolved": result.libreoffice_resolved,
            "spreadbench.recalculate.error_count": len(result.errors),
        },
    )


def _write_recalc_span(
    telemetry_context: TelemetryContext | None,
    name: str,
    *,
    start_time_unix_nano: int | None = None,
    end_time_unix_nano: int | None = None,
    status: str | dict[str, Any] = "OK",
    attributes: dict[str, Any] | None = None,
) -> None:
    if telemetry_context is None:
        return
    try:
        telemetry_context.write_span(
            name,
            event_domain="host_tool",
            start_time_unix_nano=start_time_unix_nano,
            end_time_unix_nano=end_time_unix_nano,
            status=status,
            attributes=attributes,
        )
    except Exception:
        return


def _tail(value: str | bytes | None, limit: int = 512) -> str:
    if value is None:
        return ""
    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")
    return value[-limit:]


def _tail_hash(value: str | bytes | None) -> str | None:
    tail = _tail(value)
    if not tail:
        return None
    import hashlib

    return "sha256_" + hashlib.sha256(tail.encode("utf-8")).hexdigest()
