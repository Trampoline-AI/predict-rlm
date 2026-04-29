"""Tests for the spreadsheet_rlm.tools.render pipeline."""

from __future__ import annotations

import sys
from pathlib import Path

_EXAMPLE_DIR = Path(__file__).resolve().parent.parent
if str(_EXAMPLE_DIR) not in sys.path:
    sys.path.insert(0, str(_EXAMPLE_DIR))

import pytest  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from spreadsheet_rlm.tools.recalculate import _find_libreoffice  # noqa: E402
from spreadsheet_rlm.tools.render import (  # noqa: E402
    RenderResult,
    _find_pdftoppm,
    render,
    render_to_data_uri,
)

_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"


def _build_simple_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "alpha"
    ws["B2"] = 42
    ws["A3"] = "beta"
    ws["B3"] = 17
    wb.save(path)


def _build_wide_workbook(path: Path) -> None:
    """Build a workbook whose target cells (J3:N5) are off the first print page.

    Deliberately stuffs wide headers into columns A:I so that the first printed
    page of a whole-sheet export ends before column J. Used to verify that
    cell_range rendering captures a region the full-sheet render would miss.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for col in range(1, 10):
        ws.cell(1, col).value = f"HeaderCol{col:02d}_{'X' * 20}"
    marker_text = "TARGET_MARKER_Z7Q9"
    for row in range(3, 6):
        for col in range(10, 15):
            ws.cell(row, col).value = f"{marker_text}_{row}_{col}"
    wb.save(path)


def _build_multisheet_workbook(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "Alpha"
    wb["Alpha"]["A1"] = "alpha-only"
    beta = wb.create_sheet("Beta")
    beta["A1"] = "beta-only"
    beta["B2"] = "BETA_TARGET_XYZ"
    wb.save(path)


def _has_render_tools() -> bool:
    return bool(_find_libreoffice()) and bool(_find_pdftoppm())


pytestmark = pytest.mark.skipif(
    not _has_render_tools(),
    reason="render requires soffice (LibreOffice) and pdftoppm (Poppler) on PATH",
)


def test_find_pdftoppm_locates_binary():
    assert _find_pdftoppm() is not None


def test_render_returns_png_bytes(tmp_path: Path):
    xlsx = tmp_path / "simple.xlsx"
    _build_simple_workbook(xlsx)

    result = render(xlsx)

    assert isinstance(result, RenderResult)
    assert result.source_xlsx == xlsx.resolve()
    assert len(result.png_bytes) > 0
    assert result.png_bytes.startswith(_PNG_MAGIC)


def test_render_accepts_string_path(tmp_path: Path):
    xlsx = tmp_path / "simple.xlsx"
    _build_simple_workbook(xlsx)

    result = render(str(xlsx))

    assert isinstance(result, RenderResult)
    assert result.png_bytes.startswith(_PNG_MAGIC)


def test_render_missing_file_raises(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        render(tmp_path / "does_not_exist.xlsx")


def test_render_to_data_uri_format(tmp_path: Path):
    xlsx = tmp_path / "simple.xlsx"
    _build_simple_workbook(xlsx)

    uri = render_to_data_uri(xlsx)

    assert uri.startswith("data:image/png;base64,")
    # Non-trivial payload — a blank 1x1 PNG is ~68 bytes base64; real rendered
    # content is much larger, so guard against an accidentally empty image.
    assert len(uri) > 1000


def test_render_tool_wrapper_happy_path(tmp_path: Path):
    from spreadsheet_rlm.agent.skills import render as render_tool

    xlsx = tmp_path / "simple.xlsx"
    _build_simple_workbook(xlsx)

    result = render_tool(xlsx)

    assert isinstance(result, str)
    assert result.startswith("data:image/png;base64,")


def test_render_tool_wrapper_missing_file(tmp_path: Path):
    from spreadsheet_rlm.agent.skills import render as render_tool

    result = render_tool(tmp_path / "does_not_exist.xlsx")

    assert isinstance(result, str)
    assert result.startswith("Error: file not found")


def test_render_tool_wrapper_rewrites_exceptions_as_error_strings(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    """Any exception raised by the render impl must be returned as an Error: string."""
    from spreadsheet_rlm.agent import skills as skills_module

    xlsx = tmp_path / "simple.xlsx"
    _build_simple_workbook(xlsx)

    def _boom(path, **kwargs):
        raise RuntimeError("simulated rendering failure")

    monkeypatch.setattr(skills_module, "_render_impl", _boom)

    result = skills_module.render(xlsx)

    assert isinstance(result, str)
    assert result.startswith("Error:")
    assert "simulated rendering failure" in result


def test_render_registered_on_skill():
    from spreadsheet_rlm.agent.skills import libreoffice_spreadsheet_skill

    assert "render" in libreoffice_spreadsheet_skill.tools
    assert "recalculate" in libreoffice_spreadsheet_skill.tools


def test_render_with_cell_range_differs_from_whole_sheet(tmp_path: Path):
    """Range-restricted render should produce a different image than the whole-sheet render.

    LibreOffice scales the print area to fill the page, so the ranged PNG
    isn't necessarily smaller in bytes — but its content must differ, which
    is what we actually care about: the print area took effect.
    """
    xlsx = tmp_path / "wide.xlsx"
    _build_wide_workbook(xlsx)

    whole = render(xlsx)
    ranged = render(xlsx, cell_range="J3:N5")

    assert whole.png_bytes.startswith(_PNG_MAGIC)
    assert ranged.png_bytes.startswith(_PNG_MAGIC)
    assert whole.png_bytes != ranged.png_bytes
    # Non-trivial payload — both should have meaningful rendered content
    assert len(ranged.png_bytes) > 1000


def test_render_with_sheet_qualified_cell_range(tmp_path: Path):
    xlsx = tmp_path / "multi.xlsx"
    _build_multisheet_workbook(xlsx)

    result = render(xlsx, cell_range="Beta!B2")

    assert isinstance(result, RenderResult)
    assert result.png_bytes.startswith(_PNG_MAGIC)


def test_render_with_sheet_name_kwarg(tmp_path: Path):
    xlsx = tmp_path / "multi.xlsx"
    _build_multisheet_workbook(xlsx)

    result = render(xlsx, cell_range="B2", sheet_name="Beta")

    assert isinstance(result, RenderResult)
    assert result.png_bytes.startswith(_PNG_MAGIC)


def test_render_with_bad_sheet_raises(tmp_path: Path):
    xlsx = tmp_path / "multi.xlsx"
    _build_multisheet_workbook(xlsx)

    with pytest.raises(ValueError, match="not in workbook"):
        render(xlsx, cell_range="Nonexistent!A1")


def test_render_with_empty_range_raises(tmp_path: Path):
    xlsx = tmp_path / "simple.xlsx"
    _build_simple_workbook(xlsx)

    with pytest.raises(ValueError, match="empty cell range"):
        render(xlsx, cell_range="Data!")


def test_render_to_data_uri_with_cell_range(tmp_path: Path):
    xlsx = tmp_path / "wide.xlsx"
    _build_wide_workbook(xlsx)

    uri = render_to_data_uri(xlsx, cell_range="J3:N5")

    assert uri.startswith("data:image/png;base64,")
    assert len(uri) > 1000


def test_render_tool_wrapper_forwards_cell_range(tmp_path: Path):
    from spreadsheet_rlm.agent.skills import render as render_tool

    xlsx = tmp_path / "wide.xlsx"
    _build_wide_workbook(xlsx)

    result = render_tool(xlsx, cell_range="J3:N5")

    assert isinstance(result, str)
    assert result.startswith("data:image/png;base64,")
