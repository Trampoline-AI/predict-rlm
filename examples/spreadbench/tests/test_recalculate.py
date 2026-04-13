"""Tests for the spreadsheet_rlm.recalculate pipeline."""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

_EXAMPLE_DIR = Path(__file__).resolve().parent.parent
if str(_EXAMPLE_DIR) not in sys.path:
    sys.path.insert(0, str(_EXAMPLE_DIR))

import openpyxl  # noqa: E402
import pytest  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from spreadsheet_rlm import recalculate as recalc_mod  # noqa: E402
from spreadsheet_rlm.recalculate import (  # noqa: E402
    RecalcResult,
    _count_resolved,
    _find_libreoffice,
    _formula_targets,
    _pick_winner,
    recalculate,
)


def _build_workbook_with_unresolved_formulas(path: Path) -> None:
    """Write a minimal xlsx with formulas but no cached values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = 30
    ws["B1"] = "=SUM(A1:A3)"
    ws["B2"] = "=A1*A2"
    ws["B3"] = '=IF(A1>5, "big", "small")'
    wb.save(path)


def _read_cached(path: Path, sheet: str, coord: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        return wb[sheet][coord].value
    finally:
        wb.close()


def test_formula_targets_finds_all_formula_cells(tmp_path: Path):
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)
    targets = _formula_targets(xlsx)
    assert sorted(targets) == [("Data", "B1"), ("Data", "B2"), ("Data", "B3")]


def test_count_resolved_is_zero_for_freshly_written_formulas(tmp_path: Path):
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)
    targets = _formula_targets(xlsx)
    assert _count_resolved(xlsx, targets) == 0


def test_recalculate_populates_cached_values_via_formulas_library(tmp_path: Path):
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)

    result = recalculate(xlsx)

    assert isinstance(result, RecalcResult)
    assert result.total_formulas == 3
    assert result.baseline_resolved == 0
    assert result.resolved == 3
    assert result.source == "formulas"
    assert result.formulas_resolved == 3
    assert result.libreoffice_resolved is None  # short-circuited
    assert result.errors == []

    assert _read_cached(xlsx, "Data", "B1") == 60
    assert _read_cached(xlsx, "Data", "B2") == 200
    assert _read_cached(xlsx, "Data", "B3") == "big"


def test_recalculate_preserves_non_formula_cells_and_sheet_name(tmp_path: Path):
    xlsx = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "MixedCase"
    ws["A1"] = "Header"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = 10
    ws["A3"] = 20
    ws["B1"] = "note"
    ws["B2"] = "=SUM(A2:A3)"
    wb.save(xlsx)

    recalculate(xlsx)

    wb2 = openpyxl.load_workbook(xlsx, data_only=True)
    try:
        assert wb2.sheetnames == ["MixedCase"]
        ws2 = wb2["MixedCase"]
        assert ws2["A1"].value == "Header"
        assert ws2["A2"].value == 10
        assert ws2["A3"].value == 20
        assert ws2["B1"].value == "note"
        assert ws2["B2"].value == 30
    finally:
        wb2.close()


def test_recalculate_is_noop_when_already_fully_resolved(tmp_path: Path):
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)
    recalculate(xlsx)
    before = xlsx.read_bytes()

    result = recalculate(xlsx)

    # Second call: either baseline short-circuit (no formulas left because
    # the first call flattened them) or formulas re-runs on an empty
    # target set — both should report the file as already resolved.
    assert result.source == "baseline"
    assert result.resolved == result.total_formulas
    assert xlsx.read_bytes() == before


def test_recalculate_is_strictly_additive(tmp_path: Path):
    """Running recalc twice never loses previously cached values.

    After the first pass the formulas library flattens formulas to values,
    so the second pass sees zero formula cells — but the cached values
    from the first pass must still be readable.
    """
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)

    first = recalculate(xlsx)
    assert first.resolved == first.total_formulas == 3

    b1_before = _read_cached(xlsx, "Data", "B1")
    b2_before = _read_cached(xlsx, "Data", "B2")
    b3_before = _read_cached(xlsx, "Data", "B3")

    second = recalculate(xlsx)
    assert second.source == "baseline"

    assert _read_cached(xlsx, "Data", "B1") == b1_before == 60
    assert _read_cached(xlsx, "Data", "B2") == b2_before == 200
    assert _read_cached(xlsx, "Data", "B3") == b3_before == "big"


def test_recalculate_handles_workbook_with_no_formulas(tmp_path: Path):
    xlsx = tmp_path / "w.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["A2"] = 42
    wb.save(xlsx)

    result = recalculate(xlsx)
    assert result.total_formulas == 0
    assert result.resolved == 0
    assert result.source == "baseline"
    assert result.errors == []


def test_recalculate_raises_on_missing_file(tmp_path: Path):
    with pytest.raises(FileNotFoundError):
        recalculate(tmp_path / "does_not_exist.xlsx")


def test_pick_winner_prefers_baseline_on_ties(tmp_path: Path):
    p = tmp_path / "dummy.xlsx"
    p.write_bytes(b"")
    candidates = {
        "baseline": (p, 5),
        "formulas": (p, 5),
        "libreoffice": (p, 5),
    }
    assert _pick_winner(candidates) == "baseline"


def test_pick_winner_prefers_libreoffice_over_formulas_on_ties(tmp_path: Path):
    p = tmp_path / "dummy.xlsx"
    p.write_bytes(b"")
    candidates = {
        "baseline": (p, 0),
        "formulas": (p, 5),
        "libreoffice": (p, 5),
    }
    assert _pick_winner(candidates) == "libreoffice"


def test_pick_winner_picks_strict_max(tmp_path: Path):
    p = tmp_path / "dummy.xlsx"
    p.write_bytes(b"")

    candidates = {
        "baseline": (p, 2),
        "formulas": (p, 7),
        "libreoffice": (p, 5),
    }
    assert _pick_winner(candidates) == "formulas"

    candidates = {
        "baseline": (p, 2),
        "formulas": (p, 5),
        "libreoffice": (p, 7),
    }
    assert _pick_winner(candidates) == "libreoffice"


def test_recalculate_falls_back_when_formulas_library_is_absent(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)

    monkeypatch.setattr(recalc_mod, "_formulas_lib", None)

    if _find_libreoffice() is None:
        pytest.skip("LibreOffice not installed — fallback cannot be tested")

    result = recalculate(xlsx)

    assert result.source == "libreoffice"
    assert result.resolved == 3
    assert result.formulas_resolved is None
    assert any(e.startswith("formulas:") for e in result.errors)
    assert _read_cached(xlsx, "Data", "B1") == 60
    assert _read_cached(xlsx, "Data", "B2") == 200


def test_recalculate_reports_errors_when_both_pipelines_fail(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)
    before = xlsx.read_bytes()

    monkeypatch.setattr(recalc_mod, "_formulas_lib", None)
    monkeypatch.setattr(recalc_mod, "_find_libreoffice", lambda: None)

    result = recalculate(xlsx)

    assert result.source == "baseline"
    assert result.resolved == 0
    assert any(e.startswith("formulas:") for e in result.errors)
    assert any(e.startswith("libreoffice:") for e in result.errors)
    assert xlsx.read_bytes() == before


@pytest.mark.integration
def test_recalculate_uses_libreoffice_when_formulas_is_incomplete(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    """Force the formulas library to produce zero resolved cells and verify
    LibreOffice rescues the workbook."""
    if _find_libreoffice() is None:
        pytest.skip("LibreOffice not installed")

    xlsx = tmp_path / "w.xlsx"
    _build_workbook_with_unresolved_formulas(xlsx)

    original = recalc_mod._recalc_with_formulas

    def broken(src: Path, dst: Path) -> None:
        # Simulate a formulas library that ran but produced nothing useful:
        # copy the source through unchanged so the candidate stays at 0.
        shutil.copy2(src, dst)

    monkeypatch.setattr(recalc_mod, "_recalc_with_formulas", broken)

    result = recalculate(xlsx)

    monkeypatch.setattr(recalc_mod, "_recalc_with_formulas", original)

    assert result.formulas_resolved == 0
    assert result.libreoffice_resolved == 3
    assert result.source == "libreoffice"
    assert _read_cached(xlsx, "Data", "B1") == 60
